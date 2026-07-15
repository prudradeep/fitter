from __future__ import annotations

import asyncio
import os
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from tests.generate_open_conversation_selection_test_cases import (
    COLUMNS,
    SUMMARY_SHEET,
    TEST_SHEET,
    create_workbook,
    make_test_cases,
)
from tests.run_open_conversation_selection_cases import (
    RESULT_COLUMNS,
    RESULTS_SHEET,
    run_cases,
    write_results_workbook,
)
from tests.run_open_conversation_selection_regression import main as run_regression_main


class OpenConversationSelectionTestCaseWorkbookTests(unittest.TestCase):
    def test_generator_has_broad_required_coverage(self):
        rows = make_test_cases()
        categories = {str(row["Category"]) for row in rows}

        self.assertGreaterEqual(len(rows), 80)
        self.assertLessEqual(len(rows), 160)
        self.assertIn("Country + region + sector in one message", categories)
        self.assertIn("Full natural-language flow", categories)
        self.assertIn("Conversational references", categories)

    def test_create_workbook_writes_expected_sheets_and_headers(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output = create_workbook(Path(temp_dir) / "selection_cases.xlsx")
            self.assertTrue(output.exists())

            workbook = load_workbook(output)
            self.assertEqual(workbook.sheetnames, [TEST_SHEET, SUMMARY_SHEET])

            test_sheet = workbook[TEST_SHEET]
            headers = [cell.value for cell in test_sheet[1]]
            self.assertEqual(headers, COLUMNS)
            self.assertEqual(test_sheet.freeze_panes, "A2")
            self.assertIsNotNone(test_sheet.auto_filter.ref)

            summary_sheet = workbook[SUMMARY_SHEET]
            self.assertEqual(summary_sheet["A1"].value, "Category")
            self.assertEqual(summary_sheet["B1"].value, "Number of Test Cases")

    def test_result_workbook_writes_pass_fail_columns(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            results = asyncio.run(run_cases())
            output = write_results_workbook(results, Path(temp_dir) / "selection_results.xlsx")

            self.assertTrue(output.exists())
            workbook = load_workbook(output)
            self.assertEqual(workbook.sheetnames, [RESULTS_SHEET, SUMMARY_SHEET])

            result_sheet = workbook[RESULTS_SHEET]
            headers = [cell.value for cell in result_sheet[1]]
            self.assertEqual(headers, RESULT_COLUMNS)
            self.assertEqual(result_sheet.max_row - 1, len(make_test_cases()))
            self.assertIn("Status", headers)
            self.assertIn("Reason", headers)
            status_column = headers.index("Status") + 1
            failures = [
                result_sheet.cell(row=row_index, column=status_column).value
                for row_index in range(2, result_sheet.max_row + 1)
                if result_sheet.cell(row=row_index, column=status_column).value != "Pass"
            ]
            self.assertEqual(failures, [])

    def test_combined_regression_script_runs_from_working_directory(self):
        original_cwd = Path.cwd()
        with tempfile.TemporaryDirectory() as temp_dir:
            try:
                os.chdir(temp_dir)
                run_regression_main()
                self.assertTrue((Path(temp_dir) / "open_conversation_selection_test_cases.xlsx").exists())
                self.assertTrue(
                    (Path(temp_dir) / "open_conversation_selection_test_results.xlsx").exists()
                )
            finally:
                os.chdir(original_cwd)


if __name__ == "__main__":
    unittest.main()
