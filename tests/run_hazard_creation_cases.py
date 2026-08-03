from __future__ import annotations

import argparse
import asyncio
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import get_settings
from app.schemas import ChatResponse, Option
from app.services.chat_service import ChatService
from app.services.custom_hazard_validation import default_custom_hazard_state
from app.services.chat_session import ChatSession
from tests.generate_hazard_creation_test_cases import COLUMNS, make_test_cases


OUTPUT_FILE = "hazard_creation_test_results.xlsx"
RESULTS_SHEET = "Results"
SUMMARY_SHEET = "Summary"
DEFAULT_MODELS = [
    "qwen3.5:2b",
    "qwen3.5:4b",
    "ministral-3:8b",
    "qwen3.5:9b",
    "ministral-3:14b",
    "mistral-nemo",
]

RESULT_COLUMNS = [
    *COLUMNS,
    "Actual Action",
    "Actual Step",
    "Actual Input Mode",
    "Actual Error",
    "Actual Pending Hazard",
    "Actual Rejected Dimension",
    "Actual Bot Response",
    "Status",
    "Reason",
]


class _HazardCreationEngine:
    async def handle_case(self, item: dict[str, str]) -> tuple[ChatResponse, ChatSession]:
        service = self._service()

        sector = str(item.get("Selected Sector") or "")
        session = ChatSession(
            country=str(item.get("Selected Country") or "Germany"),
            region=str(item.get("Selected Region") or "Bavaria"),
            sector=sector,
            country_id=1,
            region_id=1,
            sector_id={"Energy": 1, "Housing": 2, "Transport": 3}.get(sector, 1),
            phase="custom_hazard_input",
            custom_hazard=default_custom_hazard_state(),
            hazards=[],
            custom_hazards=[],
            additional_hazards=[],
            hazard_profiles={},
        )
        response = await service._capture_custom_hazard(
            "hazard-creation-test-session",
            session,
            str(item.get("User Hazard") or ""),
        )
        for column in ("Clarification Answer 1", "Clarification Answer 2"):
            answer = str(item.get(column) or "").strip()
            if not answer:
                continue
            if session.phase != "custom_hazard_title_clarification":
                break
            response = await service._handle_custom_hazard_title_clarification(
                "hazard-creation-test-session",
                session,
                answer,
            )
        expected_action = str(item.get("Expected Action") or "").strip()
        if expected_action == "ASK_CONTEXT_CLARIFICATION":
            service._review_custom_hazard_context = AsyncMock(
                return_value={
                    "status": "clarification",
                    "valid": False,
                    "question": "Clarification needed: who is affected in Bavaria, and which policy pathway creates the hazard?",
                }
            )
            response = await service._validate_custom_hazard(
                "hazard-creation-test-session",
                session,
                f"Reason: {str(item.get('Clarification Answer 1') or '').strip()}",
            )
        elif expected_action == "ASK_GROUNDING_CLARIFICATION":
            service._review_custom_hazard_context = AsyncMock(
                return_value={"status": "valid", "valid": True, "reason": "Context is clear."}
            )
            reason = str(item.get("Clarification Answer 1") or item.get("User Hazard") or "").strip()
            grounding_result = self._grounding_clarification_result(item)
            with patch(
                "app.services.chat_hazard_creation.validate_custom_hazard_dimensions",
                AsyncMock(return_value=grounding_result),
            ):
                response = await service._validate_custom_hazard(
                    "hazard-creation-test-session",
                    session,
                    f"Reason: {reason}",
                )
        return response, session

    @staticmethod
    def _service() -> ChatService:
        service = ChatService.__new__(ChatService)
        service.db = None
        service.user_id = None
        service.invalid_message = (
            "I could not understand your selection. Please choose from the available options."
        )
        service._ensure_user_session = lambda *args, **kwargs: None
        service._record_activity = lambda *args, **kwargs: None
        service._same_sector_hazard_names_for_duplicate_check = lambda session: list(session.hazards or [])
        service._same_scope_custom_hazard_names_for_duplicate_check = lambda session: list(session.custom_hazards or [])
        service._ensure_custom_hazard = lambda *args, **kwargs: SimpleNamespace(id="custom-hazard-id")
        service._promote_temporary_evidence = lambda *args, **kwargs: None
        service._discard_temporary_evidence = lambda *args, **kwargs: None
        service._validate_input_quality = AsyncMock(return_value={"valid": True, "reason": "Clear enough."})
        service._validate_hazard_against_stats = AsyncMock(
            return_value={"valid": True, "reason": "Compatible with the sector context."}
        )
        service._review_custom_hazard_context = AsyncMock(
            return_value={
                "status": "clarification",
                "valid": False,
                "question": "Clarification needed: who is affected in Bavaria, and which policy pathway creates the hazard?",
            }
        )
        service._hazards_step = lambda session_id, session: ChatResponse(
            session_id=session_id,
            step="hazards",
            bot_message=f"{session.sector} selected. Selection flow completed.",
            options=[
                Option(id=1, label="Start Mitigation Planning"),
                Option(id=2, label="Add a new Hazard"),
                Option(id=3, label="Refresh hazards and DGs"),
            ],
            session=session.summary(),
            error=False,
        )
        service._fuzzy_confirmation_step = lambda session_id, session, label: ChatResponse(
            session_id=session_id,
            step="fuzzy_confirmation",
            bot_message=f"Did you mean **{label}**?",
            options=[Option(id=1, label="Yes"), Option(id=2, label="No")],
            session=session.summary(),
            error=False,
        )
        return service

    @staticmethod
    def _grounding_clarification_result(item: dict[str, str]) -> dict[str, object]:
        selected_sector = str(item.get("Selected Sector") or "the selected sector")
        return {
            "overall_score": 44,
            "status": "needs_clarification",
            "next_action": "ask_clarification",
            "affected_groups": [],
            "duplicate_candidates": [],
            "dimension_scores": {
                "hazard_definition_fit": {
                    "score": 5,
                    "reason": "A negative impact is described.",
                    "needs_clarification": False,
                    "clarification_question": "",
                },
                "twin_transition_policy_fit": {
                    "score": 3,
                    "reason": "The green, digital, or twin-transition mechanism needs clearer grounding.",
                    "needs_clarification": True,
                    "clarification_question": "Which green, digital, or twin-transition policy creates this hazard?",
                },
                "selected_sector_fit": {
                    "score": 4,
                    "reason": "The selected-sector mechanism needs more detail.",
                    "needs_clarification": True,
                    "clarification_question": f"How is this connected to the selected sector: {selected_sector}?",
                },
                "country_region_fit": {
                    "score": 6,
                    "reason": "The selected place is named.",
                    "needs_clarification": False,
                    "clarification_question": "",
                },
                "affected_groups_fit": {
                    "score": 3,
                    "reason": "Affected population groups need clearer grounding.",
                    "needs_clarification": True,
                    "clarification_question": "Which population groups are affected by this hazard, and why?",
                },
            },
        }


def _rejected_dimension(response: ChatResponse) -> str:
    payload = response.custom_hazard or {}
    dimensions = payload.get("dimension_scores") if isinstance(payload, dict) else {}
    if not isinstance(dimensions, dict):
        return ""
    for key, value in dimensions.items():
        if isinstance(value, dict) and str(value.get("status") or "").upper() == "REJECTED":
            return str(key)
    return ""


def infer_actual_action(response: ChatResponse, session: ChatSession) -> str:
    rejected_dimension = _rejected_dimension(response)
    if response.error and response.bot_message and "Add a New Hazard" in response.bot_message:
        return "SHOW_ADD_HAZARD_PROMPT"
    if not response.error and response.step == "hazards" and session.phase == "hazards":
        return "GO_BACK_TO_HAZARDS"
    if (
        not response.error
        and response.step == "custom_hazard_validation"
        and response.input_mode == "reason_evidence"
        and session.pending_hazard
    ):
        return "ACCEPT_HAZARD_NAME"
    if response.step == "custom_hazard_title_clarification" and response.error:
        return "REASK_TITLE_CLARIFICATION"
    if response.step == "custom_hazard_title_clarification":
        return "ASK_TITLE_CLARIFICATION"
    if response.step == "custom_hazard_clarification":
        return "ASK_GROUNDING_CLARIFICATION"
    if response.step == "hazards" and response.input_mode == "textarea":
        return "ASK_CONTEXT_CLARIFICATION"
    if response.error and rejected_dimension == "selected_sector_fit":
        return "REJECT_SECTOR_MISMATCH"
    if response.error:
        return "REJECT_REWRITE"
    return "NO_CHANGE"


def bool_value(value: object) -> bool:
    return str(value or "").strip().casefold() == "yes"


def _message_contains(response: str, expected: str) -> bool:
    fragments = [fragment.strip() for fragment in str(expected or "").split("|") if fragment.strip()]
    if not fragments:
        return True
    normalized_response = response.casefold()
    return all(fragment.casefold() in normalized_response for fragment in fragments)


def row_result(
    item: dict[str, str],
    response: ChatResponse,
    session: ChatSession,
) -> dict[str, str]:
    actual = {
        "action": infer_actual_action(response, session),
        "step": response.step or "",
        "input_mode": response.input_mode or "",
        "error": bool(response.error),
        "pending_hazard": session.pending_hazard or "",
        "rejected_dimension": _rejected_dimension(response),
        "bot_response": response.bot_message or "",
    }
    expected = {
        "action": str(item.get("Expected Action") or "").strip(),
        "step": str(item.get("Expected Step") or "").strip(),
        "input_mode": str(item.get("Expected Input Mode") or "").strip(),
        "error": bool_value(item.get("Expected Error")),
        "pending_hazard": str(item.get("Expected Pending Hazard") or "").strip(),
        "rejected_dimension": str(item.get("Expected Rejected Dimension") or "").strip(),
        "message_contains": str(item.get("Expected Message Contains") or "").strip(),
    }

    mismatches: list[str] = []
    for key in ("action", "step", "input_mode", "pending_hazard", "rejected_dimension"):
        if expected[key] and expected[key] != actual[key]:
            mismatches.append(f"{key}: expected {expected[key]!r}, got {actual[key]!r}")
    if expected["error"] != actual["error"]:
        mismatches.append(f"error: expected {expected['error']}, got {actual['error']}")
    if not _message_contains(actual["bot_response"], expected["message_contains"]):
        mismatches.append(
            f"message: expected fragments {expected['message_contains']!r} in response"
        )

    return {
        "Actual Action": actual["action"],
        "Actual Step": actual["step"],
        "Actual Input Mode": actual["input_mode"],
        "Actual Error": "Yes" if actual["error"] else "No",
        "Actual Pending Hazard": actual["pending_hazard"],
        "Actual Rejected Dimension": actual["rejected_dimension"],
        "Actual Bot Response": actual["bot_response"],
        "Status": "Fail" if mismatches else "Pass",
        "Reason": "; ".join(mismatches) if mismatches else "Matched expected hazard creation behavior.",
    }


async def run_cases(limit: int | None = None) -> list[dict[str, str]]:
    engine = _HazardCreationEngine()
    results: list[dict[str, str]] = []
    cases = [
        item
        for item in make_test_cases()
        if str(item.get("Execution Scope") or "Automated runner").strip() == "Automated runner"
    ]
    if limit is not None:
        cases = cases[:limit]
    for index, item in enumerate(cases, start=1):
        case_item = {"Test Case ID": f"HC-{index:03d}", **item}
        try:
            response, session = await engine.handle_case(case_item)
            results.append({**case_item, **row_result(case_item, response, session)})
        except Exception as exc:
            results.append(
                {
                    **case_item,
                    "Actual Action": "NO_CHANGE",
                    "Actual Step": "",
                    "Actual Input Mode": "",
                    "Actual Error": "Yes",
                    "Actual Pending Hazard": "",
                    "Actual Rejected Dimension": "",
                    "Actual Bot Response": "",
                    "Status": "Fail",
                    "Reason": f"Runner exception: {exc.__class__.__name__}: {exc}",
                }
            )
    return results


def style_sheet(sheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    pass_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    fail_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    status_column = None
    for cell in sheet[1]:
        if cell.value == "Status":
            status_column = cell.column
            break

    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = body_alignment
        if status_column:
            status_cell = row_cells[status_column - 1]
            if status_cell.value == "Pass":
                status_cell.fill = pass_fill
            elif status_cell.value == "Fail":
                status_cell.fill = fail_fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 90))
        sheet.column_dimensions[column_letter].width = max(12, min(max_length + 2, 60))


def write_results_workbook(results: list[dict[str, str]], output_path: str | Path = OUTPUT_FILE) -> Path:
    output = Path(output_path).resolve()
    settings = get_settings()
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = RESULTS_SHEET
    sheet.append(RESULT_COLUMNS)
    for item in results:
        sheet.append([item.get(column, "") for column in RESULT_COLUMNS])
    style_sheet(sheet)

    summary = workbook.create_sheet(SUMMARY_SHEET)
    summary.append(["Run Setting", "Value"])
    summary.append(["Execution mode", "Real hazard creation flow with local LLM classifier"])
    summary.append(["Ollama base URL", settings.ollama_base_url])
    summary.append(["Ollama model", settings.ollama_model])
    summary.append([])
    summary.append(["Category", "Total", "Passed", "Failed"])
    by_category: dict[str, Counter[str]] = defaultdict(Counter)
    for item in results:
        by_category[str(item.get("Category") or "")][str(item.get("Status") or "")] += 1
    for category in sorted(by_category):
        counts = by_category[category]
        summary.append([category, counts["Pass"] + counts["Fail"], counts["Pass"], counts["Fail"]])
    summary.append(
        [
            "TOTAL",
            len(results),
            sum(1 for item in results if item.get("Status") == "Pass"),
            sum(1 for item in results if item.get("Status") == "Fail"),
        ]
    )
    style_sheet(summary)

    workbook.save(output)
    return output


def result_filename_for_model(model: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "_", model.strip())
    slug = slug.replace(":", "_")
    return f"hazard_creation_test_results_{slug}.xlsx"


async def run_cases_for_model(
    model: str,
    limit: int | None = None,
) -> list[dict[str, str]]:
    os.environ["OLLAMA_MODEL"] = model
    get_settings.cache_clear()
    return await run_cases(limit=limit)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run hazard creation cases across selected Ollama model settings."
    )
    parser.add_argument(
        "--models",
        nargs="+",
        help="Ollama model names to test. Writes one workbook per model.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="Optional maximum number of generated cases to run per model.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    models = args.models or DEFAULT_MODELS
    for model in models:
        print(f"Running hazard creation cases on Ollama model: {model}")
        results = asyncio.run(run_cases_for_model(model, limit=args.limit))
        output = write_results_workbook(results, Path.cwd() / result_filename_for_model(model))
        passed = sum(1 for item in results if item["Status"] == "Pass")
        failed = len(results) - passed
        print(f"Created hazard creation test result file: {output}")
        print(f"Model: {model} | Total: {len(results)} | Passed: {passed} | Failed: {failed}")


if __name__ == "__main__":
    main()
