from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_FILE = "open_conversation_selection_test_cases.xlsx"
TEST_SHEET = "Test Cases"
SUMMARY_SHEET = "Summary"

COUNTRIES = ["Germany", "Spain", "Portugal", "Ireland", "Italy", "Hungary"]
SECTORS = ["Energy", "Housing", "Transport"]

REGIONS = {
    "Germany": ["Bavaria", "Berlin", "Hamburg", "Hesse", "Saxony"],
    "Spain": ["Andalusia", "Catalonia", "Madrid", "Valencia", "Basque Country"],
    "Portugal": ["Lisbon", "Porto", "Alentejo", "Algarve", "Madeira"],
    "Ireland": ["Dublin", "Cork", "Galway", "Limerick", "Waterford"],
    "Italy": ["Lazio", "Lombardy", "Tuscany", "Sicily", "Veneto"],
    "Hungary": ["Budapest", "Pest", "Gyor-Moson-Sopron", "Baranya", "Csongrad-Csanad"],
}

COLUMNS = [
    "Test Case ID",
    "Category",
    "Step / Current Phase",
    "Initial State",
    "User Message",
    "Expected Country",
    "Expected Region",
    "Expected Sector",
    "Expected Hazard",
    "Expected Mitigation Measure",
    "Expected Bot Response",
    "Expected Action",
    "Should Ask Clarification",
    "Should Show Error",
    "Error Message",
    "Notes",
]


def row(
    category: str,
    phase: str,
    initial_state: str,
    message: str,
    expected_country: str = "",
    expected_region: str = "",
    expected_sector: str = "",
    expected_hazard: str = "",
    expected_mitigation_measure: str = "",
    bot_response: str = "",
    action: str = "NO_CHANGE",
    clarify: bool = False,
    error: bool = False,
    error_message: str = "",
    notes: str = "",
) -> dict[str, str | bool]:
    return {
        "Category": category,
        "Step / Current Phase": phase,
        "Initial State": initial_state,
        "User Message": message,
        "Expected Country": expected_country,
        "Expected Region": expected_region,
        "Expected Sector": expected_sector,
        "Expected Hazard": expected_hazard,
        "Expected Mitigation Measure": expected_mitigation_measure,
        "Expected Bot Response": bot_response,
        "Expected Action": action,
        "Should Ask Clarification": "Yes" if clarify else "No",
        "Should Show Error": "Yes" if error else "No",
        "Error Message": error_message,
        "Notes": notes,
    }


def make_test_cases() -> list[dict[str, str | bool]]:
    rows: list[dict[str, str | bool]] = []
    country_options = "Germany, Spain, Portugal, Ireland, Italy, Hungary"

    for country in COUNTRIES:
        rows.append(
            row(
                "Exact country selection",
                "Country",
                "No selection",
                country,
                expected_country=country,
                bot_response=f"{country} selected. Please choose a region.",
                action="SELECT_COUNTRY",
            )
        )

    natural_country_messages = [
        ("I want to start with Germany", "Germany"),
        ("Let's analyze Spain first", "Spain"),
        ("Can we look at Portugal?", "Portugal"),
        ("Start my session in Ireland please", "Ireland"),
        ("I will go with Italy", "Italy"),
        ("Set the country to Hungary", "Hungary"),
    ]
    for message, country in natural_country_messages:
        rows.append(
            row(
                "Natural-language country selection",
                "Country",
                "No selection",
                message,
                expected_country=country,
                bot_response=f"{country} selected. Please choose a region.",
                action="SELECT_COUNTRY",
            )
        )

    country_typos = [
        ("Germny", "Germany"),
        ("Spane", "Spain"),
        ("Portugul", "Portugal"),
        ("Irelnd", "Ireland"),
        ("Itlay", "Italy"),
        ("Hungry", "Hungary"),
    ]
    for message, country in country_typos:
        rows.append(
            row(
                "Country typo handling",
                "Country",
                "No selection",
                message,
                expected_country=country,
                bot_response=f"{country} selected. Please choose a region.",
                action="SELECT_COUNTRY",
                notes="Minor typo should resolve without confirmation when confidence is high.",
            )
        )

    country_abbreviations = [
        ("DE", "Germany"),
        ("Deutschland", "Germany"),
        ("ES", "Spain"),
        ("PT", "Portugal"),
        ("IE", "Ireland"),
        ("HU", "Hungary"),
    ]
    for message, country in country_abbreviations:
        rows.append(
            row(
                "Country abbreviations",
                "Country",
                "No selection",
                message,
                expected_country=country,
                bot_response=f"{country} selected. Please choose a region.",
                action="SELECT_COUNTRY",
            )
        )

    rows.extend(
        [
            row(
                "Multiple countries in one message",
                "Country",
                "No selection",
                "Germany or Spain",
                bot_response="I found multiple possible countries. Please clarify which one you want.",
                action="ASK_CLARIFICATION",
                clarify=True,
            ),
            row(
                "Multiple countries in one message",
                "Country",
                "No selection",
                "Compare Portugal and Italy",
                bot_response="I found multiple possible countries. Please clarify which one you want.",
                action="ASK_CLARIFICATION",
                clarify=True,
            ),
            row(
                "Unsupported country",
                "Country",
                "No selection",
                "France",
                bot_response=f"This country is not available. Please select one of: {country_options}.",
                action="SHOW_ERROR",
                error=True,
                error_message="Unsupported country.",
            ),
            row(
                "Random input",
                "Country",
                "No selection",
                "start something nice",
                bot_response="I could not understand your selection. Please choose from the available options.",
                action="ASK_CLARIFICATION",
                clarify=True,
            ),
        ]
    )

    for country, regions in REGIONS.items():
        for region in regions[:2]:
            rows.append(
                row(
                    "Exact region selection",
                    "Region",
                    f"Country={country}",
                    region,
                    expected_country=country,
                    expected_region=region,
                    bot_response=f"{region} selected. Please choose a sector.",
                    action="SELECT_REGION",
                )
            )

    region_natural = [
        ("Use Bavaria as the region", "Germany", "Bavaria"),
        ("I want Berlin", "Germany", "Berlin"),
        ("Let's go with Catalonia", "Spain", "Catalonia"),
        ("Choose Lisbon please", "Portugal", "Lisbon"),
        ("Set it to Dublin", "Ireland", "Dublin"),
        ("The region is Lombardy", "Italy", "Lombardy"),
        ("Budapest works for me", "Hungary", "Budapest"),
    ]
    for message, country, region_name in region_natural:
        rows.append(
            row(
                "Region natural language",
                "Region",
                f"Country={country}",
                message,
                expected_country=country,
                expected_region=region_name,
                bot_response=f"{region_name} selected. Please choose a sector.",
                action="SELECT_REGION",
            )
        )

    region_typos = [
        ("Bavria", "Germany", "Bavaria"),
        ("Cataluna", "Spain", "Catalonia"),
        ("Lisbom", "Portugal", "Lisbon"),
        ("Dubln", "Ireland", "Dublin"),
        ("Lombarddy", "Italy", "Lombardy"),
        ("Budapst", "Hungary", "Budapest"),
    ]
    for message, country, region_name in region_typos:
        rows.append(
            row(
                "Region typo handling",
                "Region",
                f"Country={country}",
                message,
                expected_country=country,
                expected_region=region_name,
                bot_response=f"{region_name} selected. Please choose a sector.",
                action="SELECT_REGION",
            )
        )

    rows.extend(
        [
            row(
                "Invalid region",
                "Region",
                "Country=Germany",
                "Queensland",
                expected_country="Germany",
                bot_response="I could not understand your selection. Please choose from the available options.",
                action="SHOW_ERROR",
                error=True,
                error_message="Invalid region.",
            ),
            row(
                "Region from wrong country",
                "Region",
                "Country=Germany",
                "Catalonia",
                expected_country="Germany",
                bot_response="Catalonia is not available for Germany. Please choose a region from Germany.",
                action="SHOW_ERROR",
                error=True,
                error_message="Region belongs to a different country.",
            ),
            row(
                "Ambiguous region",
                "Region",
                "Country=Hungary",
                "Budapest or Pest",
                expected_country="Hungary",
                bot_response="I found multiple possible regions. Please clarify which one you want.",
                action="ASK_CLARIFICATION",
                clarify=True,
                notes="Covers short names that may match more than one option or alias.",
            ),
        ]
    )

    for sector in SECTORS:
        rows.append(
            row(
                "Exact sector selection",
                "Sector",
                "Country=Germany; Region=Bavaria",
                sector,
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector=sector,
                bot_response=f"{sector} selected. Selection flow completed.",
                action="COMPLETE_SELECTION",
            )
        )

    sector_natural = [
        ("I want the Energy sector", "Energy"),
        ("Let's work on Housing", "Housing"),
        ("Transport is the one", "Transport"),
    ]
    for message, sector in sector_natural:
        rows.append(
            row(
                "Sector natural language",
                "Sector",
                "Country=Germany; Region=Bavaria",
                message,
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector=sector,
                bot_response=f"{sector} selected. Selection flow completed.",
                action="COMPLETE_SELECTION",
            )
        )

    sector_synonyms = [
        ("power and electricity", "Energy"),
        ("buildings and homes", "Housing"),
        ("mobility and public transit", "Transport"),
    ]
    for message, sector in sector_synonyms:
        rows.append(
            row(
                "Sector synonyms",
                "Sector",
                "Country=Germany; Region=Bavaria",
                message,
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector=sector,
                bot_response=f"{sector} selected. Selection flow completed.",
                action="COMPLETE_SELECTION",
            )
        )

    sector_typos = [("Energi", "Energy"), ("Housng", "Housing"), ("Trasport", "Transport")]
    for message, sector in sector_typos:
        rows.append(
            row(
                "Sector typo handling",
                "Sector",
                "Country=Germany; Region=Bavaria",
                message,
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector=sector,
                bot_response=f"{sector} selected. Selection flow completed.",
                action="COMPLETE_SELECTION",
            )
        )
    rows.append(
        row(
            "Exact country selection",
            "Sector",
            "Country=Germany; Region=Bavaria",
            "Italy",
            expected_country="Italy",
            bot_response=f"Italy selected. Please choose a region.",
            action="SELECT_COUNTRY",
        )
    )

    rows.append(
        row(
            "Invalid sector",
            "Sector",
            "Country=Germany; Region=Bavaria",
            "Healthcare",
            expected_country="Germany",
            expected_region="Bavaria",
            bot_response="I could not understand your selection. Please choose from the available options.",
            action="SHOW_ERROR",
            error=True,
            error_message="Invalid sector.",
        )
    )

    combined_cases = [
        (
            "Country + region in one message",
            "Country",
            "No selection",
            "Germany Bavaria",
            "Germany",
            "Bavaria",
            "",
            "Bavaria selected. Please choose a sector.",
            "SELECT_REGION",
        ),
        (
            "Country + sector in one message",
            "Country",
            "No selection",
            "Germany Energy",
            "Germany",
            "",
            "Energy",
            "Germany selected. Please choose a region.",
            "SELECT_COUNTRY",
        ),
        (
            "Region + sector in one message",
            "Region",
            "Country=Germany",
            "Bavaria Housing",
            "Germany",
            "Bavaria",
            "Housing",
            "Housing selected. Selection flow completed.",
            "COMPLETE_SELECTION",
        ),
        (
            "Country + region + sector in one message",
            "Country",
            "No selection",
            "Energy sector in Bavaria Germany",
            "Germany",
            "Bavaria",
            "Energy",
            "Energy selected. Selection flow completed.",
            "COMPLETE_SELECTION",
        ),
        (
            "Country + region + sector in one message",
            "Country",
            "No selection",
            "I want to start with Housing sector in Bavaria Germany",
            "Germany",
            "Bavaria",
            "Housing",
            "Housing selected. Selection flow completed.",
            "COMPLETE_SELECTION",
        ),
        (
            "Full natural-language flow",
            "Country",
            "No selection",
            "I'll go with the Transport sector in Catalonia, Spain",
            "Spain",
            "Catalonia",
            "Transport",
            "Transport selected. Selection flow completed.",
            "COMPLETE_SELECTION",
        ),
        (
            "Full natural-language flow",
            "Country",
            "No selection",
            "Please set Portugal, Lisbon, and Energy",
            "Portugal",
            "Lisbon",
            "Energy",
            "Energy selected. Selection flow completed.",
            "COMPLETE_SELECTION",
        ),
        (
            "Out-of-order input",
            "Country",
            "No selection",
            "Housing in Germany Bavaria",
            "Germany",
            "Bavaria",
            "Housing",
            "Housing selected. Selection flow completed.",
            "COMPLETE_SELECTION",
        ),
    ]
    for category, phase, state, message, country, region_name, sector, response, action in combined_cases:
        rows.append(
            row(
                category,
                phase,
                state,
                message,
                expected_country=country,
                expected_region=region_name,
                expected_sector=sector,
                bot_response=response,
                action=action,
            )
        )

    rows.extend(
        [
            row(
                "Change country",
                "Region",
                "Country=Germany",
                "Actually choose Spain",
                expected_country="Spain",
                bot_response="Spain selected. Please choose a region.",
                action="RESET_REGION_AND_SECTOR",
                notes="Changing country clears region and sector.",
            ),
            row(
                "Change region",
                "Sector",
                "Country=Germany; Region=Bavaria",
                "Change region to Berlin",
                expected_country="Germany",
                expected_region="Berlin",
                bot_response="Berlin selected. Please choose a sector.",
                action="RESET_SECTOR",
            ),
            row(
                "Change sector",
                "Completed",
                "Country=Germany; Region=Bavaria; Sector=Housing",
                "Switch to Transport",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Transport",
                bot_response="Transport selected. Selection flow completed.",
                action="SELECT_SECTOR",
            ),
            row(
                "Multiple corrections in one message",
                "Completed",
                "Country=Germany; Region=Bavaria; Sector=Housing",
                "Actually Spain, Catalonia, Transport",
                expected_country="Spain",
                expected_region="Catalonia",
                expected_sector="Transport",
                bot_response="Transport selected. Selection flow completed.",
                action="COMPLETE_SELECTION",
            ),
        ]
    )

    go_back_cases = [
        ("go back", "Sector", "Country=Germany; Region=Bavaria", "Germany", "", "", "GO_BACK"),
        ("back to country", "Region", "Country=Germany", "", "", "", "GO_BACK"),
        ("change previous step", "Sector", "Country=Spain; Region=Catalonia", "Spain", "", "", "GO_BACK"),
    ]
    for message, phase, state, country, region_name, sector, action in go_back_cases:
        rows.append(
            row(
                "Go back commands",
                phase,
                state,
                message,
                expected_country=country,
                expected_region=region_name,
                expected_sector=sector,
                bot_response="Okay, let's go back to the previous selection.",
                action=action,
            )
        )

    rows.extend(
        [
            row(
                "Explicit reset",
                "Completed",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "reset everything",
                bot_response="Selection reset. Please select a country from the available options.",
                action="RESET_ALL",
            ),
            row(
                "Confirmation",
                "Confirmation",
                "Pending selection=Germany",
                "Yes",
                expected_country="Germany",
                bot_response="Germany selected. Please choose a region.",
                action="SELECT_COUNTRY",
            ),
            row(
                "Negative confirmation",
                "Confirmation",
                "Pending selection=Germany",
                "No",
                bot_response="No problem. Please select a country from the available options.",
                action="NO_CHANGE",
            ),
        ]
    )

    odd_inputs = [
        (
            "Unsupported/multilingual input",
            "Country",
            "No selection",
            "Quiero Francia",
            "",
            "",
            "",
            f"This country is not available. Please select one of: {country_options}.",
            "SHOW_ERROR",
            False,
            True,
            "Unsupported country.",
        ),
        (
            "Unsupported/multilingual input",
            "Country",
            "No selection",
            "Deutschland bitte",
            "Germany",
            "",
            "",
            "Germany selected. Please choose a region.",
            "SELECT_COUNTRY",
            False,
            False,
            "",
        ),
        (
            "Emoji input",
            "Country",
            "No selection",
            "Germany 🇩🇪",
            "Germany",
            "",
            "",
            "Germany selected. Please choose a region.",
            "SELECT_COUNTRY",
            False,
            False,
            "",
        ),
        (
            "Emoji input",
            "Country",
            "No selection",
            "🚀🔥",
            "",
            "",
            "",
            "I could not understand your selection. Please choose from the available options.",
            "ASK_CLARIFICATION",
            True,
            False,
            "",
        ),
        (
            "Garbage input",
            "Region",
            "Country=Germany",
            "asdf qwer zxcv",
            "Germany",
            "",
            "",
            "I could not understand your selection. Please choose from the available options.",
            "ASK_CLARIFICATION",
            True,
            False,
            "",
        ),
        (
            "Empty input",
            "Country",
            "No selection",
            "",
            "",
            "",
            "",
            "Please select a country from the available options.",
            "NO_CHANGE",
            True,
            False,
            "",
        ),
    ]
    for category, phase, state, message, country, region_name, sector, response, action, clarify, error, err in odd_inputs:
        rows.append(
            row(
                category,
                phase,
                state,
                message,
                expected_country=country,
                expected_region=region_name,
                expected_sector=sector,
                bot_response=response,
                action=action,
                clarify=clarify,
                error=error,
                error_message=err,
            )
        )

    rows.extend(
        [
            row(
                "Post-sector next step",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "next step",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Please select a hazard to start mitigation planning.",
                action="START_MITIGATION_PLANNING",
            ),
            row(
                "Post-sector next step",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "continue",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Please select a hazard to start mitigation planning.",
                action="START_MITIGATION_PLANNING",
            ),
            row(
                "Post-sector next step",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "Create mitigation",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Please select a hazard to start mitigation planning.",
                action="START_MITIGATION_PLANNING",
                notes="Clear mitigation command should not require confirmation.",
            ),
            row(
                "Post-sector next step",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "Create mitigation measure",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Please select a hazard to start mitigation planning.",
                action="START_MITIGATION_PLANNING",
                notes="Post-sector mitigation creation starts the hazard selection step first.",
            ),
            row(
                "Post-sector next step",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "the first one",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Please select a hazard to start mitigation planning.",
                action="START_MITIGATION_PLANNING",
            ),
            row(
                "Post-sector add hazard",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "second one",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Describe the new hazard for Energy.",
                action="ADD_NEW_HAZARD",
            ),
            row(
                "Post-sector add hazard",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "add a hazard",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Describe the new hazard for Energy.",
                action="ADD_NEW_HAZARD",
            ),
            row(
                "Post-sector add hazard",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "Create a new hazard",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Describe the new hazard for Energy.",
                action="ADD_NEW_HAZARD",
                notes="Clear hazard creation command should not require confirmation.",
            ),
            row(
                "Post-sector add hazard",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "I want to create a new hazard",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Describe the new hazard for Energy.",
                action="ADD_NEW_HAZARD",
            ),
            row(
                "Post-sector refresh hazards",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "last one",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Hazards refreshed.",
                action="REFRESH_HAZARDS",
            ),
            row(
                "Post-sector refresh hazards",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "refresh hazards",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Hazards refreshed.",
                action="REFRESH_HAZARDS",
            ),
            row(
                "Post-sector refresh hazards",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "Update hazards list",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Hazards refreshed.",
                action="REFRESH_HAZARDS",
                notes="Clear refresh command should not require confirmation.",
            ),
            row(
                "Post-sector navigation",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "go back",
                expected_country="Germany",
                expected_region="Bavaria",
                bot_response="Bavaria selected. Please choose a sector.",
                action="RESET_SECTOR",
                notes="From post-sector actions, go back returns to sector selection.",
            ),
            row(
                "Post-sector navigation",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "select another region",
                expected_country="Germany",
                bot_response="Germany selected. Please choose a region.",
                action="RESET_REGION_AND_SECTOR",
            ),
            row(
                "Post-sector navigation",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "start over",
                bot_response="Please select a country from the available options.",
                action="RESET_ALL",
            ),
            row(
                "Post-sector change region",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "change region to Berlin",
                expected_country="Germany",
                expected_region="Berlin",
                bot_response="Berlin selected. Please choose a sector.",
                action="RESET_SECTOR",
            ),
            row(
                "Post-sector change country",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "Actually select Spain",
                expected_country="Spain",
                bot_response="Spain selected. Please choose a region.",
                action="RESET_REGION_AND_SECTOR",
            ),
            row(
                "Post-sector go back explicit",
                "Hazards",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "go back to country",
                bot_response="Please select a country from the available options.",
                action="RESET_ALL",
            ),
            row(
                "Hazard selection exact",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "Heat stress",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                bot_response="Heat stress selected. Review affected profiles.",
                action="SELECT_HAZARD",
            ),
            row(
                "Hazard selection natural language",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "I want to mitigate heat stress",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                bot_response="Heat stress selected. Review affected profiles.",
                action="SELECT_HAZARD",
            ),
            row(
                "Hazard selection natural language",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "Let's focus on Energy poverty for mitigation",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Energy poverty",
                bot_response="Energy poverty selected. Review affected profiles.",
                action="SELECT_HAZARD",
            ),
            row(
                "Hazard selection ordinal",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "the first hazard",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                bot_response="Heat stress selected. Review affected profiles.",
                action="SELECT_HAZARD",
            ),
            row(
                "Hazard selection ordinal",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "second one",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Energy poverty",
                bot_response="Energy poverty selected. Review affected profiles.",
                action="SELECT_HAZARD",
            ),
            row(
                "Hazard selection ordinal",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "last one",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Energy poverty",
                bot_response="Energy poverty selected. Review affected profiles.",
                action="SELECT_HAZARD",
            ),
            row(
                "Hazard selection typo",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "heat stres",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="<p>Did you mean <strong>Heat stress</strong>?</p>",
                action="ASK_CLARIFICATION",
                clarify=True,
            ),
            row(
                "Hazard selection invalid",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "volcano eruption",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="I could not understand your selection. Please choose from the available options.",
                action="SHOW_ERROR",
                error=True,
                error_message="Unknown hazard.",
            ),
            row(
                "Hazard selection additional hazards",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "Show hazards added by experts",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Choose one of the hazards added by experts from the selected country-sector evidence.",
                action="ASK_CLARIFICATION",
                clarify=True,
            ),
            row(
                "Hazard selection additional hazards",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "Expert-added hazard",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Expert-added hazard",
                bot_response="Expert-added hazard selected. Review affected profiles.",
                action="SELECT_HAZARD",
            ),
            row(
                "Hazard selection co-created hazards",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "Show co-created hazards",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Choose one of the co-created hazards added by users.",
                action="ASK_CLARIFICATION",
                clarify=True,
            ),
            row(
                "Hazard selection co-created hazards",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "co-created energy risk",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Co-created energy risk",
                bot_response="Co-created energy risk selected. Review affected profiles.",
                action="SELECT_HAZARD",
            ),
            row(
                "Hazard selection go back",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "go back to sectors",
                expected_country="Germany",
                expected_region="Bavaria",
                bot_response="Bavaria selected. Please choose a sector.",
                action="RESET_SECTOR",
            ),
            row(
                "Hazard selection reset",
                "Hazard",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "start over",
                bot_response="Please select a country from the available options.",
                action="RESET_ALL",
            ),
            row(
                "Reason confirmation adopt proposal",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "Adopt mitigation proposal suggested above",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                expected_mitigation_measure="Targeted heat pump support for vulnerable households",
                bot_response="Proposed mitigation measure shown with selected country, region, and sector.",
                action="ADOPT_MITIGATION_PROPOSAL",
            ),
            row(
                "Reason confirmation adopt proposal",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "show the proposed mitigation measure",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                expected_mitigation_measure="Targeted heat pump support for vulnerable households",
                bot_response="Proposed mitigation measure shown with selected country, region, and sector.",
                action="ADOPT_MITIGATION_PROPOSAL",
            ),
            row(
                "Reason confirmation adopt proposal",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "use the suggested proposal",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                expected_mitigation_measure="Targeted heat pump support for vulnerable households",
                bot_response="Proposed mitigation measure shown with selected country, region, and sector.",
                action="ADOPT_MITIGATION_PROPOSAL",
            ),
            row(
                "Reason confirmation manual mitigation",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "Yes",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                bot_response="Please share the mitigation measure you would recommend.",
                action="WRITE_MITIGATION_MANUALLY",
            ),
            row(
                "Reason confirmation manual mitigation",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "write my own",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                bot_response="Please share the mitigation measure you would recommend.",
                action="WRITE_MITIGATION_MANUALLY",
            ),
            row(
                "Reason confirmation ordinal",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "second one",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                bot_response="Other actions shown.",
                action="NO_CHANGE",
            ),
            row(
                "Reason confirmation ordinal",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "2nd one",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                bot_response="Other actions shown.",
                action="NO_CHANGE",
            ),
            row(
                "Reason confirmation ordinal",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "last one",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                expected_mitigation_measure="Targeted heat pump support for vulnerable households",
                bot_response="Proposed mitigation measure shown with selected country, region, and sector.",
                action="ADOPT_MITIGATION_PROPOSAL",
            ),
            row(
                "Reason confirmation ordinal",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "2nd last",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                bot_response="Other actions shown.",
                action="NO_CHANGE",
            ),
            row(
                "Reason confirmation change sector",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "change sector to Housing",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Housing",
                bot_response="Housing selected. Selection flow completed.",
                action="COMPLETE_SELECTION",
                notes="Navigation/change commands must not be captured as mitigation measures.",
            ),
            row(
                "Reason confirmation go back",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "change previous step",
                expected_country="Germany",
                expected_region="Bavaria",
                bot_response="Bavaria selected. Please choose a sector.",
                action="RESET_SECTOR",
            ),
            row(
                "Reason confirmation typed mitigation",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "Provide targeted cooling grants for low-income households exposed to heat stress.",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                expected_mitigation_measure="Provide targeted cooling grants for low-income households exposed to heat stress.",
                bot_response="Typed mitigation measure accepted and reason requested.",
                action="CAPTURE_MITIGATION_MEASURE",
            ),
            row(
                "Reason confirmation skip",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "not now",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                bot_response="Other actions shown.",
                action="NO_CHANGE",
            ),
            row(
                "Reason confirmation invalid",
                "Reason confirmation",
                "Country=Germany; Region=Bavaria; Sector=Energy; Hazard=Heat stress",
                "maybe",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                expected_hazard="Heat stress",
                bot_response="I could not understand your selection. Please choose from the available options.",
                action="SHOW_ERROR",
                error=True,
                error_message="Unclear reason confirmation response.",
            ),
            row(
                "Long descriptive input",
                "Country",
                "No selection",
                "For this assessment, I want to focus on the Bavarian housing transition context in Germany.",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Housing",
                bot_response="Housing selected. Selection flow completed.",
                action="COMPLETE_SELECTION",
            ),
            row(
                "Invalid country-sector combination",
                "Country",
                "No selection",
                "France Energy",
                bot_response=f"This country is not available. Please select one of: {country_options}.",
                action="SHOW_ERROR",
                error=True,
                error_message="Unsupported country.",
            ),
            row(
                "Repeated selection",
                "Region",
                "Country=Germany",
                "Germany again",
                expected_country="Germany",
                bot_response="Germany is already selected. Please choose a region.",
                action="NO_CHANGE",
            ),
            row(
                "Repeated selection",
                "Completed",
                "Country=Germany; Region=Bavaria; Sector=Energy",
                "Energy",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Energy",
                bot_response="Energy is already selected. Selection flow completed.",
                action="NO_CHANGE",
            ),
            row(
                "Conversational references",
                "Region",
                "Country=Germany",
                "the first region option",
                expected_country="Germany",
                expected_region="Bavaria",
                bot_response="Bavaria selected. Please choose a sector.",
                action="SELECT_REGION",
            ),
            row(
                "Conversational references",
                "Country",
                "No selection",
                "the last one",
                expected_country="Spain",
                bot_response="Spain selected. Please choose a region.",
                action="SELECT_COUNTRY",
            ),
            row(
                "Conversational references",
                "Country",
                "No selection",
                "2nd one",
                expected_country="Hungary",
                bot_response="Hungary selected. Please choose a region.",
                action="SELECT_COUNTRY",
            ),
            row(
                "Conversational references",
                "Country",
                "No selection",
                "third one",
                expected_country="Ireland",
                bot_response="Ireland selected. Please choose a region.",
                action="SELECT_COUNTRY",
            ),
            row(
                "Conversational references",
                "Country",
                "No selection",
                "2nd last",
                expected_country="Portugal",
                bot_response="Portugal selected. Please choose a region.",
                action="SELECT_COUNTRY",
            ),
            row(
                "Conversational references",
                "Country",
                "No selection",
                "3rd last",
                expected_country="Italy",
                bot_response="Italy selected. Please choose a region.",
                action="SELECT_COUNTRY",
            ),
            row(
                "Conversational references",
                "Region",
                "Country=Germany",
                "second one",
                expected_country="Germany",
                expected_region="Berlin",
                bot_response="Berlin selected. Please choose a sector.",
                action="SELECT_REGION",
            ),
            row(
                "Conversational references",
                "Region",
                "Country=Germany",
                "last one",
                expected_country="Germany",
                expected_region="Saxony",
                bot_response="Saxony selected. Please choose a sector.",
                action="SELECT_REGION",
            ),
            row(
                "Conversational references",
                "Region",
                "Country=Germany",
                "2nd last",
                expected_country="Germany",
                expected_region="Hesse",
                bot_response="Hesse selected. Please choose a sector.",
                action="SELECT_REGION",
            ),
            row(
                "Conversational references",
                "Sector",
                "Country=Germany; Region=Bavaria",
                "the last one",
                expected_country="Germany",
                expected_region="Bavaria",
                expected_sector="Transport",
                bot_response="Transport selected. Selection flow completed.",
                action="COMPLETE_SELECTION",
            ),
        ]
    )

    return rows


def add_test_cases_sheet(workbook: Workbook, rows: list[dict[str, str | bool]]) -> None:
    sheet = workbook.active
    sheet.title = TEST_SHEET
    sheet.append(COLUMNS)

    for index, item in enumerate(rows, start=1):
        sheet.append([f"TC-{index:03d}", *[item.get(column, "") for column in COLUMNS[1:]]])

    style_sheet(sheet)


def add_summary_sheet(workbook: Workbook, rows: list[dict[str, str | bool]]) -> None:
    sheet = workbook.create_sheet(SUMMARY_SHEET)
    sheet.append(["Category", "Number of Test Cases"])
    counts = Counter(str(item["Category"]) for item in rows)
    for category in sorted(counts):
        sheet.append([category, counts[category]])
    style_sheet(sheet)


def style_sheet(sheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = body_alignment

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
        sheet.column_dimensions[column_letter].width = max(14, min(max_length + 2, 55))


def create_workbook(output_path: str | Path = OUTPUT_FILE) -> Path:
    output = Path(output_path).resolve()
    rows = make_test_cases()

    workbook = Workbook()
    add_test_cases_sheet(workbook, rows)
    add_summary_sheet(workbook, rows)
    workbook.save(output)
    return output


def main() -> None:
    output = create_workbook(Path.cwd() / OUTPUT_FILE)
    print(f"Created Excel test-case file: {output}")


if __name__ == "__main__":
    main()
