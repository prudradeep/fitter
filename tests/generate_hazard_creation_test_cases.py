from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_FILE = "hazard_creation_test_cases.xlsx"
TEST_SHEET = "Test Cases"
SUMMARY_SHEET = "Summary"

COUNTRY = "Germany"
REGION = "Bavaria"
SECTORS = ["Energy", "Housing", "Transport"]

COLUMNS = [
    "Test Case ID",
    "Category",
    "Selected Country",
    "Selected Region",
    "Selected Sector",
    "User Hazard",
    "Clarification Answer 1",
    "Clarification Answer 2",
    "Reason / Justification",
    "Evidence Decision",
    "Evidence Input",
    "Context Clarification Answer",
    "Grounding Clarification Answer",
    "Affected Group Review Action",
    "Affected Group Reason",
    "Strict Validation",
    "Crowd Sourcing",
    "Execution Scope",
    "Expected Action",
    "Expected Step",
    "Expected Input Mode",
    "Expected Error",
    "Expected Pending Hazard",
    "Expected Rejected Dimension",
    "Expected Message Contains",
    "Notes",
]


VALID_HAZARDS = {
    "Energy": [
        "Low-income households face higher electricity bills from renewable grid upgrade tariffs",
        "Rural residents face power outages from grid congestion during renewable energy integration",
        "Small businesses face utility arrears from dynamic electricity pricing and smart meter rollout",
        "Tenants face higher clean heating bills as heat pump tariffs and grid upgrade charges increase",
        "Energy community members face exclusion from solar sharing because smart grid access fees rise",
        "Fossil-dependent regional workers face job losses and local tax-base decline from coal power phase-out",
    ],
    "Housing": [
        "Low-income tenants face rent increases after mandatory green building retrofits",
        "Apartment residents face temporary displacement during deep renovation and insulation mandates",
        "Homeowners in inefficient buildings face unaffordable compliance costs from energy performance rules",
        "Renters face eviction pressure when landlords pass building renovation costs into higher rents",
        "Older residents in poorly insulated homes face disruption during mandatory residential retrofit works",
        "Low-income tenants face renovation cost burden and renoviction from stricter building energy performance standards",
    ],
    "Transport": [
        "Low-income commuters face higher travel costs from low emission zone fees",
        "Rural residents lose mobility access when bus routes are replaced by digital booking systems",
        "Taxi drivers face income loss from EV charging downtime and clean vehicle mandates",
        "Apartment dwellers lose access to EV adoption because transport electrification relies on home charging",
        "Delivery workers face income loss from clean vehicle rules and scarce public charging infrastructure",
    ],
}

INVALID_HAZARDS = [
    "I like sunny weather",
    "Tell me about retrofit policy",
    "Solar panels reduce electricity bills for everyone",
    "Please add more data to the chart",
    "Carbon monoxide poisoning from domestic heating",
    "Structural housing hazards from missing smoke detectors and window guards",
]

VAGUE_OR_INCOMPLETE_HAZARDS = [
    "The green transition is complicated for residents",
    "Digitalisation may affect some people",
    "Energy policy changes are happening quickly",
    "Housing renovation is important",
    "Transport electrification is a major topic",
]

BENEFIT_OR_MITIGATION_STATEMENTS = [
    "Solar subsidies reduce energy bills for low-income households",
    "Retrofit grants help tenants live in warmer homes",
    "EV charging grants support taxi drivers",
    "Better bus services improve access for rural residents",
    "Smart meters help households understand energy use",
]

CROSS_SECTOR_HAZARDS = {
    "Energy": "Low-income households face rising electricity tariffs from renewable grid upgrades",
    "Housing": "Tenants face rent increases after mandatory apartment retrofit policies",
    "Transport": "Taxi drivers face income loss from EV charging downtime and clean vehicle mandates",
}

SECTOR_SYNONYM_HAZARDS = {
    "Energy": [
        "Households face utility arrears when power tariffs rise after smart grid modernisation",
        "Remote communities face outage risk from renewable power balancing constraints",
    ],
    "Housing": [
        "Renters face unaffordable dwelling upgrades from stricter residential energy performance rules",
        "Flat residents face disruption when landlords renovate buildings to meet insulation standards",
    ],
    "Transport": [
        "Commuters face mobility exclusion when public transit ticketing becomes digital-only",
        "Drivers without private parking face barriers to clean vehicle adoption from charging access gaps",
    ],
}

MIXED_SIGNAL_VALID_HAZARDS = {
    "Energy": [
        "Apartment households face higher electricity bills when clean heating tariffs fund grid upgrades",
        "Tenants face utility arrears because heat pump electricity tariffs rise faster than incomes",
    ],
    "Housing": [
        "Households face rent increases when landlords retrofit homes and install heat pumps",
        "Apartment residents face displacement during building renovation for energy efficiency compliance",
    ],
    "Transport": [
        "Renters face unequal EV adoption because transport electrification depends on private home charging",
        "Low-income households face mobility exclusion when clean vehicle zones raise car access costs",
    ],
}

TITLE_CLARIFICATION_FLOWS = [
    {
        "sector": "Energy",
        "hazard": "Digital energy services leave people behind.",
        "answer_1": "",
        "answer_2": "",
        "expected_action": "ASK_TITLE_CLARIFICATION",
        "expected_step": "custom_hazard_title_clarification",
        "expected_input_mode": "text",
        "expected_error": False,
        "expected_message_contains": "Clarification Needed",
        "notes": "A transition-linked but underspecified hazard title should ask a title clarification question.",
    },
    {
        "sector": "Energy",
        "hazard": "Digital energy services leave people behind.",
        "answer_1": "I don't know",
        "answer_2": "",
        "expected_action": "REASK_TITLE_CLARIFICATION",
        "expected_step": "custom_hazard_title_clarification",
        "expected_input_mode": "text",
        "expected_error": True,
        "expected_message_contains": "does not clarify the hazard|Clarification Needed",
        "notes": "A non-answer must not be accepted as title clarification.",
    },
    {
        "sector": "Energy",
        "hazard": "Digital energy services leave people behind.",
        "answer_1": "Older adults are affected, but I am not sure how.",
        "answer_2": "what to do",
        "expected_action": "REASK_TITLE_CLARIFICATION",
        "expected_step": "custom_hazard_title_clarification",
        "expected_input_mode": "text",
        "expected_error": True,
        "expected_message_contains": "question or request|Clarification Needed",
        "notes": "Question/request-style replies must be rejected in later title clarification rounds too.",
    },
    {
        "sector": "Energy",
        "hazard": "Digital energy services leave people behind.",
        "answer_1": "Older adults and low-income households without internet access or digital skills are excluded from online-only electricity billing and support services.",
        "answer_2": "",
        "expected_action": "ACCEPT_HAZARD_NAME",
        "expected_step": "custom_hazard_validation",
        "expected_input_mode": "reason_evidence",
        "expected_error": False,
        "expected_message_contains": "Reason and Evidence Needed",
        "notes": "A meaningful clarification should be validated together with the original hazard title and proceed to reason/evidence.",
    },
]

CONTEXT_CLARIFICATION_FLOWS = [
    {
        "sector": "Energy",
        "hazard": "Low-income households face higher electricity bills from renewable grid upgrade tariffs",
        "reason": "This affects households in Germany, but I have not explained the Bavarian policy or tariff pathway yet.",
        "expected_action": "ASK_CONTEXT_CLARIFICATION",
        "expected_step": "hazards",
        "expected_input_mode": "textarea",
        "expected_error": False,
        "expected_message_contains": "Clarification|Bavaria",
        "notes": "After reason/evidence, locally incomplete support should ask a context clarification before saving the hazard.",
    },
    {
        "sector": "Housing",
        "hazard": "Tenants face rent increases after mandatory green building retrofits",
        "reason": "Retrofit obligations can raise rents, but the explanation does not identify whether tenants or owners bear the cost.",
        "expected_action": "ASK_CONTEXT_CLARIFICATION",
        "expected_step": "hazards",
        "expected_input_mode": "textarea",
        "expected_error": False,
        "expected_message_contains": "Clarification|who is affected",
        "notes": "A plausible hazard with an unresolved affected-group mechanism should stay in clarification rather than being rejected or accepted.",
    },
]

GROUNDING_DIMENSION_CLARIFICATION_FLOWS = [
    {
        "sector": "Transport",
        "hazard": "Clean mobility access problems",
        "answer_1": "Low-income commuters in Bavaria face higher travel costs when low-emission zone rules restrict older cars before affordable electric alternatives are available.",
        "expected_action": "ASK_GROUNDING_CLARIFICATION",
        "expected_step": "custom_hazard_clarification",
        "expected_input_mode": "textarea",
        "expected_error": False,
        "expected_message_contains": "more detail|selected sector",
        "notes": "A clarified title can still need dimension grounding when the affected group or policy pathway is not grounded enough.",
    },
    {
        "sector": "Energy",
        "hazard": "Smart meter rollout creates cost pressure",
        "answer_1": "Low-income households in Bavaria face installation and tariff costs when smart meter rollout shifts billing and grid charges onto consumers.",
        "expected_action": "ASK_GROUNDING_CLARIFICATION",
        "expected_step": "custom_hazard_clarification",
        "expected_input_mode": "textarea",
        "expected_error": False,
        "expected_message_contains": "more detail|green, digital",
        "notes": "Grounding clarification should be distinct from title clarification and occur after the hazard title has been accepted.",
    },
]

FULL_FLOW_SCENARIOS = [
    {
        "category": "Full flow - reason to evidence decision",
        "sector": "Energy",
        "hazard": "Low-income households face higher electricity bills from renewable grid upgrade tariffs",
        "reason": "Bavarian low-income households can face higher electricity bills when renewable grid upgrade costs are passed through tariffs before income support adjusts.",
        "expected_action": "ASK_EVIDENCE_DECISION",
        "expected_step": "custom_hazard_evidence_decision",
        "expected_input_mode": "",
        "expected_message_contains": "Do you have evidence|Yes|No",
        "notes": "After a valid title and reason, the current flow asks whether the user has optional evidence.",
    },
    {
        "category": "Full flow - open no evidence",
        "sector": "Housing",
        "hazard": "Low-income tenants face rent increases after mandatory green building retrofits",
        "reason": "In Bavaria, stricter building energy standards can push retrofit costs into rents for low-income tenants before protections or subsidies offset the increase.",
        "evidence_decision": "no, I don't have evidence",
        "expected_action": "VALIDATE_WITHOUT_EVIDENCE",
        "expected_step": "custom_hazard_group_review",
        "expected_input_mode": "",
        "expected_message_contains": "Hazard to be co-created",
        "notes": "Open-chat no-evidence replies should proceed to validation without being blocked by input-quality checks.",
    },
    {
        "category": "Full flow - evidence input requested",
        "sector": "Transport",
        "hazard": "Low-income commuters face higher travel costs from low emission zone fees",
        "reason": "Low-income commuters in Bavaria who depend on older cars can face higher travel costs when low-emission zone fees are introduced before affordable alternatives are available.",
        "evidence_decision": "yes, I have evidence",
        "expected_action": "ASK_EVIDENCE_INPUT",
        "expected_step": "custom_hazard_evidence",
        "expected_input_mode": "evidence_only",
        "expected_message_contains": "Paste a URL|supported file|Skip",
        "notes": "A Yes evidence decision should enter the evidence-only input step.",
    },
    {
        "category": "Full flow - URL evidence in open chat",
        "sector": "Energy",
        "hazard": "Fossil-dependent regional workers face job losses and local tax-base decline from coal power phase-out",
        "reason": "Energy transition policies that phase out coal power can concentrate job losses and local revenue decline in fossil-dependent regions while new clean-energy benefits arrive elsewhere.",
        "evidence_decision": "Use this evidence https://example.org/coal-transition-report.pdf",
        "evidence_input": "https://example.org/coal-transition-report.pdf",
        "expected_action": "VALIDATE_WITH_URL_EVIDENCE",
        "expected_step": "custom_hazard_group_review",
        "expected_input_mode": "",
        "expected_message_contains": "Hazard to be co-created",
        "notes": "A URL embedded in open-chat evidence-decision text should be normalized as evidence and validated.",
    },
    {
        "category": "Full flow - evidence skip",
        "sector": "Housing",
        "hazard": "Low-income tenants face renovation cost burden and renoviction from stricter building energy performance standards",
        "reason": "Stricter building energy performance standards can trigger renovation costs and eviction pressure for low-income tenants if landlords recover costs through rent hikes or vacant-possession strategies.",
        "evidence_decision": "Yes",
        "evidence_input": "Skip",
        "expected_action": "VALIDATE_WITHOUT_EVIDENCE",
        "expected_step": "custom_hazard_group_review",
        "expected_input_mode": "",
        "expected_message_contains": "Hazard to be co-created",
        "notes": "Choosing Skip at evidence input should validate the staged hazard without evidence.",
    },
    {
        "category": "Full flow - evidence contradiction",
        "sector": "Transport",
        "hazard": "Taxi drivers face income loss from EV charging downtime and clean vehicle mandates",
        "reason": "Taxi drivers can lose working time and income if clean vehicle mandates require EV adoption before reliable public fast charging is available.",
        "evidence_decision": "Yes",
        "evidence_input": "Evidence states that fast charging is already universal and taxi operating time increased after the mandate.",
        "expected_action": "REJECT_EVIDENCE_CONTRADICTION",
        "expected_step": "custom_hazard_validation",
        "expected_error": True,
        "expected_rejected_dimension": "evidence_support",
        "expected_message_contains": "contradict|evidence",
        "notes": "Evidence that contradicts the hazard should reject rather than save or clarify.",
    },
    {
        "category": "Full flow - reason quality rejection",
        "sector": "Energy",
        "hazard": "Tenants face utility arrears because heat pump electricity tariffs rise faster than incomes",
        "reason": "because it is bad",
        "expected_action": "REJECT_REASON_QUALITY",
        "expected_step": "custom_hazard_validation",
        "expected_error": True,
        "expected_message_contains": "reason|justification",
        "notes": "A low-quality reason should be rejected before evidence, context, or group review.",
    },
    {
        "category": "Full flow - reason reveals sector mismatch",
        "sector": "Housing",
        "hazard": "Households face rising costs from clean transition policies",
        "reason": "The real issue is low-emission zone fees and vehicle access restrictions that make car commutes more expensive.",
        "expected_action": "REJECT_SECTOR_MISMATCH_AFTER_REASON",
        "expected_step": "custom_hazard_validation",
        "expected_error": True,
        "expected_rejected_dimension": "selected_sector_fit",
        "expected_message_contains": "sector",
        "notes": "Reason/evidence text can reveal that a short title belongs to another selected sector.",
    },
    {
        "category": "Full flow - context clarification resolved",
        "sector": "Energy",
        "hazard": "Smart meter rollout creates cost pressure",
        "reason": "Smart meter rollout can shift installation and billing costs onto consumers.",
        "context_answer": "In Bavaria, low-income households in rented apartments are affected when smart-meter installation fees and time-of-use tariffs are passed into electricity bills before targeted subsidies are available.",
        "expected_action": "RESOLVE_CONTEXT_CLARIFICATION",
        "expected_step": "custom_hazard_group_review",
        "expected_input_mode": "",
        "expected_message_contains": "Hazard to be co-created",
        "notes": "A useful answer to the context-review question should continue validation instead of restarting title clarification.",
    },
    {
        "category": "Full flow - duplicate confirmation",
        "sector": "Transport",
        "hazard": "Taxi drivers face income loss from EV charging downtime and clean vehicle mandates",
        "reason": "Taxi drivers can lose fare time when clean vehicle mandates arrive before reliable public charging is available.",
        "expected_action": "ASK_DUPLICATE_CONFIRMATION",
        "expected_step": "custom_hazard_duplicate_confirmation",
        "expected_message_contains": "Possible Duplicate Hazard|Continue with custom hazard",
        "notes": "Near-duplicate custom hazards should ask for confirmation rather than looping or silently saving.",
    },
    {
        "category": "Full flow - duplicate override",
        "sector": "Transport",
        "hazard": "Taxi drivers face income loss from EV charging downtime and clean vehicle mandates",
        "reason": "Taxi drivers can lose fare time when clean vehicle mandates arrive before reliable public charging is available.",
        "affected_group_action": "Continue with custom hazard",
        "expected_action": "CONTINUE_AFTER_DUPLICATE_OVERRIDE",
        "expected_step": "custom_hazard_group_review",
        "expected_message_contains": "Hazard to be co-created",
        "notes": "Choosing to continue with a custom hazard should set duplicate_override_confirmed and move forward.",
    },
    {
        "category": "Full flow - duplicate use existing",
        "sector": "Transport",
        "hazard": "Taxi drivers face income loss from EV charging downtime and clean vehicle mandates",
        "reason": "Taxi drivers can lose fare time when clean vehicle mandates arrive before reliable public charging is available.",
        "affected_group_action": "Use existing hazard",
        "expected_action": "SELECT_EXISTING_DUPLICATE_HAZARD",
        "expected_step": "custom_hazard_population_review",
        "expected_message_contains": "socio-demographic",
        "notes": "Choosing the suggested existing hazard should select it and leave custom hazard creation.",
    },
    {
        "category": "Full flow - grounding clarification resolved",
        "sector": "Transport",
        "hazard": "Clean mobility access problems",
        "reason": "Low-income commuters can face travel-cost increases from low-emission zone fees.",
        "grounding_answer": "This is a transport-sector hazard because low-emission zone rules restrict access by older cars, and low-income commuters in Bavaria lack affordable EVs or replacement public transport.",
        "expected_action": "RESOLVE_GROUNDING_CLARIFICATION",
        "expected_step": "custom_hazard_group_review",
        "expected_message_contains": "Hazard to be co-created",
        "notes": "Grounding answers should be combined with the original reason/evidence and then route to group review when resolved.",
    },
    {
        "category": "Full flow - generic affected group rejected",
        "sector": "Energy",
        "hazard": "Low-income households face higher electricity bills from renewable grid upgrade tariffs",
        "reason": "People are affected by higher electricity costs from grid upgrades.",
        "affected_group_action": "Add affected group: people",
        "expected_action": "REASK_AFFECTED_GROUP",
        "expected_step": "custom_hazard_group_review",
        "expected_error": True,
        "expected_message_contains": "specific affected group",
        "notes": "Generic groups such as people, households, residents, consumers, and general population should not be accepted.",
    },
    {
        "category": "Full flow - add affected group asks reason",
        "sector": "Housing",
        "hazard": "Renters face eviction pressure when landlords pass building renovation costs into higher rents",
        "reason": "Landlords may pass retrofit costs into rents, increasing eviction pressure for renters with limited income.",
        "affected_group_action": "Add affected group: pensioners on fixed incomes",
        "expected_action": "ASK_AFFECTED_GROUP_REASON",
        "expected_step": "custom_hazard_profile_reason",
        "expected_input_mode": "textarea",
        "expected_message_contains": "How does this hazard affect",
        "notes": "When a user adds a specific affected group without a reason, the app should ask for the impact reason.",
    },
    {
        "category": "Full flow - added affected group reason accepted",
        "sector": "Housing",
        "hazard": "Renters face eviction pressure when landlords pass building renovation costs into higher rents",
        "reason": "Landlords may pass retrofit costs into rents, increasing eviction pressure for renters with limited income.",
        "affected_group_action": "Add affected group: pensioners on fixed incomes",
        "affected_group_reason": "They have limited ability to absorb rent increases after renovation costs are passed through.",
        "expected_action": "ACCEPT_AFFECTED_GROUP_REASON",
        "expected_step": "custom_hazard_group_review",
        "expected_message_contains": "pensioners on fixed incomes",
        "notes": "A clear group-impact reason should add the group and return to group review.",
    },
    {
        "category": "Full flow - edit affected group reason",
        "sector": "Energy",
        "hazard": "Tenants face utility arrears because heat pump electricity tariffs rise faster than incomes",
        "reason": "Heat pump tariff increases can push low-income tenants into arrears when electricity bills rise faster than incomes.",
        "affected_group_action": "low-income tenants: tariff increases create arrears risk when incomes are fixed",
        "expected_action": "EDIT_AFFECTED_GROUP_REASON",
        "expected_step": "custom_hazard_group_review",
        "expected_message_contains": "low-income tenants",
        "notes": "Colon-style group edits should update the matching affected-group reason.",
    },
    {
        "category": "Full flow - remove user added group",
        "sector": "Transport",
        "hazard": "Low-income commuters face higher travel costs from low emission zone fees",
        "reason": "Low-emission zone fees can increase travel costs for commuters who cannot replace older cars quickly.",
        "affected_group_action": "Remove group: students without cars",
        "expected_action": "REMOVE_USER_ADDED_GROUP",
        "expected_step": "custom_hazard_group_review",
        "expected_message_contains": "Hazard to be co-created",
        "notes": "User-added groups can be removed during review.",
    },
    {
        "category": "Full flow - system group removal blocked",
        "sector": "Transport",
        "hazard": "Low-income commuters face higher travel costs from low emission zone fees",
        "reason": "Low-emission zone fees can increase travel costs for commuters who cannot replace older cars quickly.",
        "affected_group_action": "Remove group: low-income commuters",
        "expected_action": "BLOCK_SYSTEM_GROUP_REMOVAL",
        "expected_step": "custom_hazard_group_review",
        "expected_error": True,
        "expected_message_contains": "can't be removed|system",
        "notes": "System-identified affected groups cannot be removed; users can add or edit instead.",
    },
    {
        "category": "Full flow - confirm affected groups",
        "sector": "Energy",
        "hazard": "Low-income households face higher electricity bills from renewable grid upgrade tariffs",
        "reason": "Renewable grid upgrade tariffs can increase electricity bills for low-income households before support is available.",
        "affected_group_action": "Confirm affected groups",
        "expected_action": "SAVE_CUSTOM_HAZARD",
        "expected_step": "custom_hazard_population_review",
        "expected_message_contains": "co-created hazard",
        "notes": "Confirming reviewed groups should save the custom hazard and continue to the custom hazard population step.",
    },
    {
        "category": "Full flow - no profiles target population",
        "sector": "Energy",
        "hazard": "Small businesses face utility arrears from dynamic electricity pricing and smart meter rollout",
        "reason": "Dynamic electricity pricing can raise billing volatility for small businesses that cannot shift demand.",
        "expected_action": "ASK_TARGET_POPULATION",
        "expected_step": "target_population_question",
        "expected_input_mode": "textarea",
        "expected_message_contains": "target population",
        "notes": "If no affected profiles are extracted, the app should ask target-population questions before review.",
    },
    {
        "category": "Full flow - strict crowd review notice",
        "sector": "Housing",
        "hazard": "Low-income tenants face rent increases after mandatory green building retrofits",
        "reason": "Retrofit costs can be passed into rents for low-income tenants.",
        "strict_validation": "Yes",
        "crowd_sourcing": "Yes",
        "expected_action": "SHOW_REVIEW_VISIBILITY_NOTICE",
        "expected_step": "custom_hazard_group_review",
        "expected_message_contains": "platform users|Bavaria|Germany",
        "notes": "Strict validation with Crowd Sourcing should show the platform-visibility notice on group review.",
    },
    {
        "category": "Full flow - strict crowd success notice",
        "sector": "Housing",
        "hazard": "Low-income tenants face rent increases after mandatory green building retrofits",
        "reason": "Retrofit costs can be passed into rents for low-income tenants.",
        "affected_group_action": "Confirm affected groups",
        "strict_validation": "Yes",
        "crowd_sourcing": "Yes",
        "expected_action": "SHOW_SUCCESS_VISIBILITY_NOTICE",
        "expected_step": "custom_hazard_population_review",
        "expected_message_contains": "visible to platform users|Bavaria|Germany",
        "notes": "Strict validation with Crowd Sourcing should also show visibility text on final success.",
    },
]


def row(
    *,
    category: str,
    sector: str,
    hazard: str,
    expected_action: str,
    expected_step: str,
    clarification_answer_1: str = "",
    clarification_answer_2: str = "",
    reason: str = "",
    evidence_decision: str = "",
    evidence_input: str = "",
    context_answer: str = "",
    grounding_answer: str = "",
    affected_group_action: str = "",
    affected_group_reason: str = "",
    strict_validation: str = "",
    crowd_sourcing: str = "",
    execution_scope: str = "Automated runner",
    expected_input_mode: str = "",
    expected_error: bool = False,
    expected_pending_hazard: str = "",
    expected_rejected_dimension: str = "",
    expected_message_contains: str = "",
    notes: str = "",
) -> dict[str, str]:
    return {
        "Category": category,
        "Selected Country": COUNTRY,
        "Selected Region": REGION,
        "Selected Sector": sector,
        "User Hazard": hazard,
        "Clarification Answer 1": clarification_answer_1,
        "Clarification Answer 2": clarification_answer_2,
        "Reason / Justification": reason,
        "Evidence Decision": evidence_decision,
        "Evidence Input": evidence_input,
        "Context Clarification Answer": context_answer,
        "Grounding Clarification Answer": grounding_answer,
        "Affected Group Review Action": affected_group_action,
        "Affected Group Reason": affected_group_reason,
        "Strict Validation": strict_validation,
        "Crowd Sourcing": crowd_sourcing,
        "Execution Scope": execution_scope,
        "Expected Action": expected_action,
        "Expected Step": expected_step,
        "Expected Input Mode": expected_input_mode,
        "Expected Error": "Yes" if expected_error else "No",
        "Expected Pending Hazard": expected_pending_hazard,
        "Expected Rejected Dimension": expected_rejected_dimension,
        "Expected Message Contains": expected_message_contains,
        "Notes": notes,
    }


def make_test_cases() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []

    for sector, hazards in VALID_HAZARDS.items():
        for hazard in hazards:
            rows.append(
                row(
                    category="Valid hazard for selected sector",
                    sector=sector,
                    hazard=hazard,
                    expected_action="ACCEPT_HAZARD_NAME",
                    expected_step="custom_hazard_validation",
                    expected_input_mode="reason_evidence",
                    expected_pending_hazard=hazard,
                    expected_message_contains="Reason and Evidence Needed",
                    notes="A clear sector-specific transition risk should proceed to reason/evidence collection.",
                )
            )

    for sector, hazards in SECTOR_SYNONYM_HAZARDS.items():
        for hazard in hazards:
            rows.append(
                row(
                    category="Valid hazard with sector synonyms",
                    sector=sector,
                    hazard=hazard,
                    expected_action="ACCEPT_HAZARD_NAME",
                    expected_step="custom_hazard_validation",
                    expected_input_mode="reason_evidence",
                    expected_pending_hazard=hazard,
                    expected_message_contains="Reason and Evidence Needed",
                    notes="Synonyms such as utility, dwelling, flat, transit, or mobility should still fit the selected sector.",
                )
            )

    for sector, hazards in MIXED_SIGNAL_VALID_HAZARDS.items():
        for hazard in hazards:
            rows.append(
                row(
                    category="Valid mixed-signal hazard",
                    sector=sector,
                    hazard=hazard,
                    expected_action="ACCEPT_HAZARD_NAME",
                    expected_step="custom_hazard_validation",
                    expected_input_mode="reason_evidence",
                    expected_pending_hazard=hazard,
                    expected_message_contains="Reason and Evidence Needed",
                    notes="Some hazards mention adjacent-sector terms but should pass when the selected-sector mechanism is dominant.",
                )
            )

    for sector in SECTORS:
        for hazard in INVALID_HAZARDS:
            is_general_safety = "Carbon monoxide" in hazard
            is_structural_safety = "Structural housing hazards" in hazard
            rows.append(
                row(
                    category="Invalid or non-hazard input",
                    sector=sector,
                    hazard=hazard,
                    expected_action="REJECT_REWRITE",
                    expected_step="hazards",
                    expected_error=True,
                    expected_rejected_dimension="twin_transition_policy_fit",
                    expected_message_contains=(
                        "general household safety risk|please rewrite"
                        if is_general_safety or is_structural_safety
                        else "hazard"
                    ),
                    notes="Non-hazards, generic questions, benefits, or transition-unrelated risks should not continue.",
                )
            )

    for sector in SECTORS:
        rows.append(
            row(
                category="Generic consumer price issue",
                sector=sector,
                hazard="Rising grocery prices reduce household purchasing power in Germany's Baden-Württemberg region.",
                expected_action="REJECT_REWRITE",
                expected_step="hazards",
                expected_error=True,
                expected_rejected_dimension="twin_transition_policy_fit",
                expected_message_contains="grocery or food-price pressure",
                notes="General grocery-price or purchasing-power harms should not pass as sector transition hazards.",
            )
        )

    for sector in SECTORS:
        for hazard in VAGUE_OR_INCOMPLETE_HAZARDS:
            rows.append(
                row(
                    category="Vague or incomplete hazard",
                    sector=sector,
                    hazard=hazard,
                    expected_action="ASK_TITLE_CLARIFICATION",
                    expected_step="custom_hazard_title_clarification",
                    expected_input_mode="text",
                    expected_message_contains="Clarification Needed",
                    notes="Transition-linked but underspecified policy-topic statements should ask for hazard-title clarification before reason/evidence.",
                )
            )

    for flow in TITLE_CLARIFICATION_FLOWS:
        rows.append(
            row(
                category="Hazard title clarification flow",
                sector=str(flow["sector"]),
                hazard=str(flow["hazard"]),
                clarification_answer_1=str(flow["answer_1"]),
                clarification_answer_2=str(flow["answer_2"]),
                expected_action=str(flow["expected_action"]),
                expected_step=str(flow["expected_step"]),
                expected_input_mode=str(flow["expected_input_mode"]),
                expected_error=bool(flow["expected_error"]),
                expected_message_contains=str(flow["expected_message_contains"]),
                notes=str(flow["notes"]),
            )
        )

    for flow in CONTEXT_CLARIFICATION_FLOWS:
        rows.append(
            row(
                category="Hazard clarification after reason",
                sector=str(flow["sector"]),
                hazard=str(flow["hazard"]),
                clarification_answer_1=str(flow["reason"]),
                expected_action=str(flow["expected_action"]),
                expected_step=str(flow["expected_step"]),
                expected_input_mode=str(flow["expected_input_mode"]),
                expected_error=bool(flow["expected_error"]),
                expected_message_contains=str(flow["expected_message_contains"]),
                notes=str(flow["notes"]),
            )
        )

    for flow in GROUNDING_DIMENSION_CLARIFICATION_FLOWS:
        rows.append(
            row(
                category="Grounding dimension clarification flow",
                sector=str(flow["sector"]),
                hazard=str(flow["hazard"]),
                clarification_answer_1=str(flow["answer_1"]),
                expected_action=str(flow["expected_action"]),
                expected_step=str(flow["expected_step"]),
                expected_input_mode=str(flow["expected_input_mode"]),
                expected_error=bool(flow["expected_error"]),
                expected_message_contains=str(flow["expected_message_contains"]),
                notes=str(flow["notes"]),
            )
        )

    for scenario in FULL_FLOW_SCENARIOS:
        rows.append(
            row(
                category=str(scenario["category"]),
                sector=str(scenario["sector"]),
                hazard=str(scenario["hazard"]),
                reason=str(scenario.get("reason") or ""),
                evidence_decision=str(scenario.get("evidence_decision") or ""),
                evidence_input=str(scenario.get("evidence_input") or ""),
                context_answer=str(scenario.get("context_answer") or ""),
                grounding_answer=str(scenario.get("grounding_answer") or ""),
                affected_group_action=str(scenario.get("affected_group_action") or ""),
                affected_group_reason=str(scenario.get("affected_group_reason") or ""),
                strict_validation=str(scenario.get("strict_validation") or ""),
                crowd_sourcing=str(scenario.get("crowd_sourcing") or ""),
                execution_scope="Spec only",
                expected_action=str(scenario["expected_action"]),
                expected_step=str(scenario["expected_step"]),
                expected_input_mode=str(scenario.get("expected_input_mode") or ""),
                expected_error=bool(scenario.get("expected_error", False)),
                expected_rejected_dimension=str(scenario.get("expected_rejected_dimension") or ""),
                expected_message_contains=str(scenario.get("expected_message_contains") or ""),
                notes=str(scenario["notes"]),
            )
        )

    for sector in SECTORS:
        for hazard in BENEFIT_OR_MITIGATION_STATEMENTS:
            rows.append(
                row(
                    category="Benefit or mitigation statement",
                    sector=sector,
                    hazard=hazard,
                    expected_action="REJECT_REWRITE",
                    expected_step="hazards",
                    expected_error=True,
                    expected_rejected_dimension="twin_transition_policy_fit",
                    expected_message_contains="hazard",
                    notes="Benefits and mitigation actions are not hazards unless they describe a concrete harm or risk.",
                )
            )

    for selected_sector in SECTORS:
        for source_sector, hazard in CROSS_SECTOR_HAZARDS.items():
            if source_sector == selected_sector:
                continue
            rows.append(
                row(
                    category="Wrong-sector hazard",
                    sector=selected_sector,
                    hazard=hazard,
                    expected_action="REJECT_SECTOR_MISMATCH",
                    expected_step="hazards",
                    expected_error=True,
                    expected_rejected_dimension="selected_sector_fit",
                    expected_message_contains="sector",
                    notes=(
                        f"A {source_sector} hazard submitted while {selected_sector} is selected "
                        "should ask the user to rewrite it or choose the matching sector."
                    ),
                )
            )

    for sector in SECTORS:
        rows.append(
            row(
                category="Empty hazard input",
                sector=sector,
                hazard="",
                expected_action="SHOW_ADD_HAZARD_PROMPT",
                expected_step="hazards",
                expected_error=True,
                expected_message_contains=f"Add a New Hazard|{sector}",
                notes="Blank input should keep the user on hazard creation.",
            )
        )
        rows.append(
            row(
                category="Go back from hazard creation",
                sector=sector,
                hazard="Go back to list of hazards",
                expected_action="GO_BACK_TO_HAZARDS",
                expected_step="hazards",
                expected_message_contains=f"{sector} selected. Selection flow completed.",
                notes="The explicit back option should return to the post-sector hazards screen.",
            )
        )

    return rows


def add_test_cases_sheet(workbook: Workbook, rows: list[dict[str, str]]) -> None:
    sheet = workbook.active
    sheet.title = TEST_SHEET
    sheet.append(COLUMNS)

    for index, item in enumerate(rows, start=1):
        sheet.append([f"HC-{index:03d}", *[item.get(column, "") for column in COLUMNS[1:]]])

    style_sheet(sheet)


def add_summary_sheet(workbook: Workbook, rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet(SUMMARY_SHEET)
    sheet.append(["Category", "Number of Test Cases"])
    counts = Counter(str(item["Category"]) for item in rows)
    for category in sorted(counts):
        sheet.append([category, counts[category]])
    style_sheet(sheet)


def style_sheet(sheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = body_alignment

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
        sheet.column_dimensions[column_letter].width = max(14, min(max_length + 2, 55))


def create_workbook(output_path: str | Path = OUTPUT_FILE) -> Path:
    output = Path(output_path).resolve()
    rows = make_test_cases()

    workbook = Workbook()
    add_test_cases_sheet(workbook, rows)
    add_summary_sheet(workbook, rows)
    workbook.save(output)
    return output


def main() -> None:
    output = create_workbook(Path.cwd() / OUTPUT_FILE)
    print(f"Created Excel hazard-creation test-case file: {output}")


if __name__ == "__main__":
    main()
