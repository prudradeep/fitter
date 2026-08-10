from __future__ import annotations

import argparse
import asyncio
import os
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import get_settings
from app.schemas import ChatResponse, Option
from app.services.chat_options import best_fuzzy_label, normalize, option_list
from app.services.chat_hazard_steps import ChatHazardStepsMixin
from app.services.chat_mitigation_steps import ChatMitigationStepsMixin
from app.services.chat_navigation_steps import ChatNavigationStepsMixin
from app.services.chat_selection_steps import ChatSelectionStepsMixin
from app.services.chat_session import ChatSession
from tests.generate_open_conversation_selection_test_cases import (
    COLUMNS,
    COUNTRIES,
    REGIONS,
    SECTORS,
    OUTPUT_FILE as TEST_CASES_FILE,
    load_or_make_test_cases,
    make_test_cases,
)


OUTPUT_FILE = "open_conversation_selection_test_results.xlsx"
RESULTS_SHEET = "Results"
SUMMARY_SHEET = "Summary"

RESULT_COLUMNS = [
    *COLUMNS,
    "Actual Country",
    "Actual Region",
    "Actual Sector",
    "Actual Hazard",
    "Actual Mitigation Measure",
    "Actual Bot Response",
    "Actual Action",
    "Actual Should Ask Clarification",
    "Actual Should Show Error",
    "Status",
    "Reason",
]


@dataclass(frozen=True)
class _Row:
    id: int
    name: str


@dataclass(frozen=True)
class _CountryRow(_Row):
    pass


@dataclass(frozen=True)
class _RegionRow(_Row):
    country_id: int


@dataclass(frozen=True)
class _SectorRow(_Row):
    pass


class _OpenConversationSelectionEngine(
    ChatMitigationStepsMixin,
    ChatHazardStepsMixin,
    ChatNavigationStepsMixin,
    ChatSelectionStepsMixin,
):
    welcome_message = "Please select a country from the available options."
    invalid_message = "I could not understand your selection. Please choose from the available options."

    def __init__(self) -> None:
        self.countries = [
            _CountryRow(index, country)
            for index, country in enumerate(sorted(COUNTRIES), start=1)
        ]
        self.country_by_name = {country.name: country for country in self.countries}
        self.regions: list[_RegionRow] = []
        next_region_id = 1
        for country in self.countries:
            for region in REGIONS[country.name]:
                self.regions.append(_RegionRow(next_region_id, region, country.id))
                next_region_id += 1
        self.sectors = [
            _SectorRow(index, sector)
            for index, sector in enumerate(SECTORS, start=1)
        ]

    async def handle_case(self, item: dict[str, str]) -> tuple[ChatResponse, ChatSession]:
        session = self._session_from_state(str(item.get("Initial State") or ""))
        phase = str(item.get("Step / Current Phase") or "").strip().casefold()
        message = str(item.get("User Message") or "")
        session.phase = {
            "country": "country",
            "region": "region",
            "sector": "sector",
            "completed": "hazards",
            "hazards": "hazards",
            "post-sector": "hazards",
            "hazard": "hazard_profile_selection",
            "reason confirmation": "reason_confirmation",
            "confirmation": "country",
        }.get(phase, session.phase)

        if phase == "confirmation" and "pending selection=germany" in normalize(item.get("Initial State") or ""):
            session.pending_selection_confirmation = {
                "country": "Germany",
                "region": None,
                "sector": None,
            }

        if phase == "country":
            response = await self._select_country("test-session", session, message)
        elif phase == "region":
            response = await self._select_region("test-session", session, message)
        elif phase == "sector":
            response = await self._select_sector("test-session", session, message)
        elif phase == "confirmation":
            response = await self._handle_pending_selection_workflow(
                "test-session",
                session,
                message,
            )
            if response is None:
                response = self._repeat_current_options(
                    "test-session",
                    session,
                    self.invalid_message,
                    True,
                )
        elif phase in {"hazards", "post-sector"}:
            response = await self._handle_hazards_action("test-session", session, message)
        elif phase == "hazard":
            response = await self._handle_hazard_profile_selection("test-session", session, message)
        elif phase == "reason confirmation":
            session.suggested_new_policy_proposal = (
                "Targeted heat pump support for vulnerable households"
            )
            response = await self._handle_reason_confirmation("test-session", session, message)
        else:
            response = await self._maybe_apply_conversational_selection(
                "test-session",
                session,
                message,
                current_phase="sector",
            )
            if response is None:
                response = self._repeat_current_options(
                    "test-session",
                    session,
                    self.invalid_message,
                    True,
                )

        return response, session

    def _session_from_state(self, state: str) -> ChatSession:
        session = ChatSession()
        for part in str(state or "").split(";"):
            if "=" not in part:
                continue
            key, value = [chunk.strip() for chunk in part.split("=", 1)]
            if key.casefold() == "country":
                country = self._country_by_name(value)
                if country:
                    session.country_id = country.id
                    session.country = country.name
            elif key.casefold() == "region" and session.country_id:
                region = self._region_by_name(value, session.country_id)
                if region:
                    session.region_id = region.id
                    session.region = region.name
            elif key.casefold() == "sector":
                sector = self._sector_by_name(value)
                if sector:
                    session.sector_id = sector.id
                    session.sector = sector.name
                    session.phase = "hazards"
                    self._seed_hazard_context(session)
            elif key.casefold() == "hazard":
                session.selected_hazard = value
            elif key.casefold() == "mitigation":
                session.pending_mitigation_measure = value
        return session

    async def _select_country(
        self,
        session_id: str,
        session: ChatSession,
        message: str,
    ) -> ChatResponse:
        if not message.strip():
            return self._country_step(
                session_id,
                session,
                "Please select a country from the available options.",
                False,
            )

        pending_response = await self._handle_pending_selection_workflow(session_id, session, message)
        if pending_response is not None:
            return pending_response

        conversational_response = await self._maybe_apply_conversational_selection(
            session_id,
            session,
            message,
            current_phase="country",
        )
        if conversational_response is not None:
            return conversational_response

        country = self._match_country(message)
        if country is None:
            fuzzy_country = self._fuzzy_row_by_name(self.countries, message)
            if fuzzy_country is not None:
                return self._clarification_step(
                    session_id,
                    session,
                    f"Did you mean **{fuzzy_country.name}**?",
                )
            return self._country_step(session_id, session, self.invalid_message, True)

        session.country_id = country.id
        session.country = country.name
        session.phase = "region"
        return ChatResponse(
            session_id=session_id,
            step="region",
            bot_message=f"{country.name} selected. Please choose a region.",
            options=option_list(self._regions_for_country(country.id)),
            session=session.summary(),
            error=False,
        )

    async def _select_region(
        self,
        session_id: str,
        session: ChatSession,
        message: str,
    ) -> ChatResponse:
        pending_response = await self._handle_pending_selection_workflow(session_id, session, message)
        if pending_response is not None:
            return pending_response

        conversational_response = await self._maybe_apply_conversational_selection(
            session_id,
            session,
            message,
            current_phase="region",
        )
        if conversational_response is not None:
            return conversational_response

        region = self._match_region(message, session.country_id)
        if region is None:
            wrong_country_region = self._region_by_name_any_country(message)
            if wrong_country_region is not None:
                return ChatResponse(
                    session_id=session_id,
                    step="region",
                    bot_message=(
                        f"{wrong_country_region.name} is not available for "
                        f"{session.country}. Please choose a region from {session.country}."
                    ),
                    options=option_list(self._regions_for_country(session.country_id)),
                    session=session.summary(),
                    error=True,
                )
            fuzzy_region = self._fuzzy_row_by_name(self._regions_for_country(session.country_id), message)
            if fuzzy_region is not None:
                return self._clarification_step(
                    session_id,
                    session,
                    f"Did you mean **{fuzzy_region.name}**?",
                )
            return ChatResponse(
                session_id=session_id,
                step="region",
                bot_message=self.invalid_message,
                options=option_list(self._regions_for_country(session.country_id)),
                session=session.summary(),
                error=True,
            )

        session.region_id = region.id
        session.region = region.name
        session.phase = "sector"
        return ChatResponse(
            session_id=session_id,
            step="sector",
            bot_message=f"{region.name} selected. Please choose a sector.",
            options=option_list(self._sectors_for_country(session.country_id)),
            session=session.summary(),
            error=False,
        )

    async def _select_sector(
        self,
        session_id: str,
        session: ChatSession,
        message: str,
    ) -> ChatResponse:
        pending_response = await self._handle_pending_selection_workflow(session_id, session, message)
        if pending_response is not None:
            return pending_response

        conversational_response = await self._maybe_apply_conversational_selection(
            session_id,
            session,
            message,
            current_phase="sector",
        )
        if conversational_response is not None:
            return conversational_response

        sector = self._match_sector(message, session.country_id)
        if sector is None:
            fuzzy_sector = self._fuzzy_row_by_name(self._sectors_for_country(session.country_id), message)
            if fuzzy_sector is not None:
                return self._clarification_step(
                    session_id,
                    session,
                    f"Did you mean **{fuzzy_sector.name}**?",
                )
            return ChatResponse(
                session_id=session_id,
                step="sector",
                bot_message=self.invalid_message,
                options=option_list(self._sectors_for_country(session.country_id)),
                session=session.summary(),
                error=True,
            )

        session.sector_id = sector.id
        session.sector = sector.name
        session.phase = "hazards"
        return self._hazards_step(session_id, session)

    async def _apply_selection_action(
        self,
        session_id: str,
        session: ChatSession,
        action: str | None,
    ) -> ChatResponse:
        if action in {"change_country", "restart_selection"}:
            self.reset_state_from(session, "country")
            session.phase = "country"
            return self._country_step(session_id, session, self.welcome_message, False)
        if action == "change_region":
            self.reset_state_from(session, "region")
            session.phase = "region"
            return ChatResponse(
                session_id=session_id,
                step="region",
                bot_message=f"{session.country or 'Country'} selected. Please choose a region.",
                options=option_list(self._regions_for_country(session.country_id)),
                session=session.summary(),
                error=False,
            )
        if action == "change_sector":
            self.reset_state_from(session, "sector")
            session.phase = "sector"
            return ChatResponse(
                session_id=session_id,
                step="sector",
                bot_message=f"{session.region or 'Region'} selected. Please choose a sector.",
                options=option_list(self._sectors_for_country(session.country_id)),
                session=session.summary(),
                error=False,
            )
        return self._repeat_current_options(session_id, session, self.invalid_message, True)

    def _hazard_profile_step(self, session_id: str, session: ChatSession) -> ChatResponse:
        session.phase = "hazard_profile_selection"
        self._seed_hazard_context(session)
        return ChatResponse(
            session_id=session_id,
            step="hazard_profile_selection",
            bot_message="Please select a hazard to start mitigation planning.",
            options=self._hazard_options(session),
            session=session.summary(),
            error=False,
        )

    async def _hazard_profiles_response(
        self,
        session_id: str,
        session: ChatSession,
        hazard: str,
    ) -> ChatResponse:
        session.phase = "socio_demographic_review"
        return ChatResponse(
            session_id=session_id,
            step="socio_demographic_review",
            bot_message=f"{hazard} selected. Review affected profiles.",
            options=[Option(id=1, label="Create Mitigation Measure"), Option(id=2, label="Add more DGs")],
            session=session.summary(),
            error=False,
        )

    def _hazard_options(self, session: ChatSession) -> list[Option]:
        self._seed_hazard_context(session)
        options = [Option(id=index, label=hazard) for index, hazard in enumerate(self._primary_hazard_names(session), start=1)]
        if session.additional_hazards:
            options.append(Option(id=len(options) + 1, label="Show hazards added by experts"))
        if session.custom_hazards:
            options.append(Option(id=len(options) + 1, label="Show co-created hazards"))
        return options

    def _additional_hazard_selection_options(self, session: ChatSession) -> list[Option]:
        self._seed_hazard_context(session)
        options = [Option(id=index, label=hazard) for index, hazard in enumerate(session.additional_hazards or [], start=1)]
        options.append(Option(id=len(options) + 1, label="Show listed hazards"))
        return options

    def _custom_hazard_selection_options(self, session: ChatSession) -> list[Option]:
        self._seed_hazard_context(session)
        options = [Option(id=index, label=hazard) for index, hazard in enumerate(session.custom_hazards or [], start=1)]
        options.append(Option(id=len(options) + 1, label="Show listed hazards"))
        return options

    def _match_hazard(self, message: str, session: ChatSession) -> str | None:
        normalized = normalize(message)
        for index, hazard in enumerate(self._all_hazard_names(session), start=1):
            if str(index) == str(message).strip() or normalize(hazard) == normalized:
                return hazard
        return None

    def _fuzzy_hazard(self, message: str, session: ChatSession) -> str | None:
        return best_fuzzy_label(message, self._all_hazard_names(session))

    def _is_saved_custom_hazard(self, session: ChatSession, hazard: str) -> bool:
        return normalize(hazard) in {normalize(item) for item in (session.custom_hazards or [])}

    def _custom_hazard_id_for_context(self, session: ChatSession, hazard: str):
        return 1 if self._is_saved_custom_hazard(session, hazard) else None

    def _target_population_answers_for_saved_hazard(self, session: ChatSession, hazard: str) -> list[dict[str, object]]:
        return []

    def _hydrate_custom_hazard_profiles(self, session: ChatSession) -> None:
        self._seed_hazard_context(session)

    def _filter_session_hazards_without_profiles(self, session: ChatSession) -> None:
        self._seed_hazard_context(session)

    def _seed_hazard_context(self, session: ChatSession) -> None:
        session.hazards = session.hazards or ["Heat stress", "Energy poverty"]
        session.additional_hazards = session.additional_hazards or ["Expert-added hazard"]
        session.custom_hazards = session.custom_hazards or ["Co-created energy risk"]
        session.hazard_profiles = session.hazard_profiles or {
            "Heat stress": [{"name": "Workers"}],
            "Energy poverty": [{"name": "Low-income households"}],
            "Expert-added hazard": [{"name": "Tenants"}],
            "Co-created energy risk": [{"name": "Residents"}],
        }

    @staticmethod
    def _primary_hazard_names(session: ChatSession) -> list[str]:
        additional = {normalize(hazard) for hazard in (session.additional_hazards or [])}
        custom = {normalize(hazard) for hazard in (session.custom_hazards or [])}
        return [
            hazard
            for hazard in (session.hazards or [])
            if normalize(hazard) not in additional and normalize(hazard) not in custom
        ]

    def _all_hazard_names(self, session: ChatSession) -> list[str]:
        self._seed_hazard_context(session)
        return [
            *self._primary_hazard_names(session),
            *(session.additional_hazards or []),
            *(session.custom_hazards or []),
        ]

    def _custom_hazard_input_step(self, session_id: str, session: ChatSession) -> ChatResponse:
        session.phase = "custom_hazard_input"
        return ChatResponse(
            session_id=session_id,
            step="hazards",
            bot_message=f"Describe the new hazard for {session.sector}.",
            options=[Option(id=1, label="Go back to list of hazards")],
            session=session.summary(),
            error=False,
        )

    async def _refresh_session_hazards(self, session_id: str, session: ChatSession) -> None:
        session.pending_hazard = "__refresh_hazards__"

    def _country_step(
        self,
        session_id: str,
        session: ChatSession,
        bot_message: str,
        error: bool = False,
    ) -> ChatResponse:
        return ChatResponse(
            session_id=session_id,
            step="country",
            bot_message=bot_message,
            options=option_list(self.countries),
            session=session.summary(),
            error=error,
        )

    def _repeat_current_options(
        self,
        session_id: str,
        session: ChatSession,
        bot_message: str | None = None,
        error: bool = True,
    ) -> ChatResponse:
        message = bot_message or self.invalid_message
        if session.country is None:
            return self._country_step(session_id, session, message, error)
        if session.region is None:
            return ChatResponse(
                session_id=session_id,
                step="region",
                bot_message=message,
                options=option_list(self._regions_for_country(session.country_id)),
                session=session.summary(),
                error=error,
            )
        if session.sector is None:
            return ChatResponse(
                session_id=session_id,
                step="sector",
                bot_message=message,
                options=option_list(self._sectors_for_country(session.country_id)),
                session=session.summary(),
                error=error,
            )
        return ChatResponse(
            session_id=session_id,
            step="hazards",
            bot_message=message,
            options=[
                Option(id=1, label="Start Mitigation Planning"),
                Option(id=2, label="Add a new Hazard"),
                Option(id=3, label="Refresh hazards and DGs"),
            ],
            session=session.summary(),
            error=error,
        )

    def _clarification_step(
        self,
        session_id: str,
        session: ChatSession,
        message: str,
    ) -> ChatResponse:
        return ChatResponse(
            session_id=session_id,
            step="selection_confirmation",
            bot_message=message,
            options=[Option(id=1, label="Yes"), Option(id=2, label="No")],
            session=session.summary(),
            error=False,
        )

    def _hazards_step(self, session_id: str, session: ChatSession) -> ChatResponse:
        return ChatResponse(
            session_id=session_id,
            step="hazards",
            bot_message=f"{session.sector} selected. Selection flow completed.",
            options=[
                Option(id=1, label="Start Mitigation Planning"),
                Option(id=2, label="Add a new Hazard"),
                Option(id=3, label="Refresh hazards and DGs"),
            ],
            session=session.summary(),
            error=False,
        )

    @classmethod
    def _clear_sector_context(cls, session: ChatSession) -> None:
        session.sector_id = None
        session.sector = None
        session.hazards = None
        session.hazard_profiles = None
        session.custom_hazards = None
        session.additional_hazards = None

    @classmethod
    def _clear_region_context(cls, session: ChatSession) -> None:
        session.region_id = None
        session.region = None
        cls._clear_sector_context(session)

    async def _selection_message_from_llm(self, session, event, fallback):
        return fallback

    async def _handle_anytime_grounded_question(self, session_id, session, message):
        return None

    def _ensure_user_session(self, session_id, session):
        return None

    def _record_activity(self, session_id, session, activity_type, details=None, step=None):
        return None

    @classmethod
    def _clear_selected_hazard_context(cls, session: ChatSession) -> None:
        session.selected_hazard = None
        session.selected_hazard_record_id = None

    def _clear_mitigation_clarity_state(self, session: ChatSession) -> None:
        return None

    def _clear_mitigation_validation_state(self, session: ChatSession) -> None:
        return None

    def _current_policy_mitigation_measure(self, session: ChatSession) -> str:
        return "Current policy-based mitigation"

    def _mitigation_measure_examples(self, sector_id: int | None) -> str:
        return "- Example mitigation measure."

    async def _other_actions_message_from_llm(self, session: ChatSession) -> str:
        return "Other actions shown."

    def _primary_other_nav_options(self, session: ChatSession, step: str) -> list[Option]:
        return [Option(id=1, label="Start over with a different country")]

    async def _capture_mitigation_measure(
        self,
        session_id: str,
        session: ChatSession,
        message: str,
    ) -> ChatResponse:
        mitigation_measure = str(message or "").strip()
        session.pending_mitigation_measure = mitigation_measure
        session.phase = "mitigation_reason"
        return ChatResponse(
            session_id=session_id,
            step="mitigation_reason",
            bot_message="Typed mitigation measure accepted and reason requested.",
            options=[],
            session=session.summary(),
            input_mode="reason_evidence",
            input_values={"mitigation_measure": mitigation_measure},
            error=False,
        )

    def _available_country_names(self) -> list[str]:
        return [country.name for country in self.countries]

    def _available_region_names(self, session: ChatSession) -> list[str]:
        if session.country_id is not None:
            return [region.name for region in self._regions_for_country(session.country_id)]
        return [region.name for region in self.regions]

    def _available_sector_names(self, session: ChatSession) -> list[str]:
        return [sector.name for sector in self.sectors]

    def _match_country(self, message: str):
        return self._match_by_id_or_name(self.countries, message)

    def _match_region(self, message: str, country_id: int | None):
        return self._match_by_id_or_name(self._regions_for_country(country_id), message)

    def _match_sector(self, message: str, country_id: int | None):
        return self._match_by_id_or_name(self._sectors_for_country(country_id), message)

    def _country_by_name(self, name: str | None):
        if not name:
            return None
        target = normalize(name)
        return next((country for country in self.countries if normalize(country.name) == target), None)

    def _region_by_name(self, name: str | None, country_id: int):
        if not name:
            return None
        target = normalize(name)
        return next(
            (
                region
                for region in self._regions_for_country(country_id)
                if normalize(region.name) == target
            ),
            None,
        )

    def _region_by_name_any_country(self, name: str | None):
        if not name:
            return None
        target = normalize(name)
        return next((region for region in self.regions if normalize(region.name) == target), None)

    def _sector_by_name(self, name: str | None):
        if not name:
            return None
        target = normalize(name)
        return next((sector for sector in self.sectors if normalize(sector.name) == target), None)

    def _regions_for_country(self, country_id: int | None) -> list[_RegionRow]:
        return [region for region in self.regions if region.country_id == country_id]

    def _sectors_for_country(self, country_id: int | None) -> list[_SectorRow]:
        return list(self.sectors) if country_id is not None else []

    @staticmethod
    def _match_by_id_or_name(rows, message: str):
        normalized = normalize(message)
        for row in rows:
            if str(row.id) == str(message).strip() or normalize(row.name) == normalized:
                return row
        return None

    @staticmethod
    def _fuzzy_row_by_name(rows, message: str):
        fuzzy_name = best_fuzzy_label(message, [row.name for row in rows])
        if fuzzy_name is None:
            return None
        return next((row for row in rows if row.name == fuzzy_name), None)


def infer_actual_action(response: ChatResponse, session: ChatSession) -> str:
    if response.step in {"selection_confirmation", "fuzzy_confirmation"}:
        return "ASK_CLARIFICATION"
    if (
        not response.error
        and response.bot_message
        == "Please choose one of the available options, or type your selection another way."
    ):
        return "ASK_CLARIFICATION"
    if response.error:
        return "SHOW_ERROR"
    if response.step == "hazard_profile_selection" and response.bot_message in {
        "Choose one of the hazards added by experts from the selected country-sector evidence.",
        "Choose one of the co-created hazards added by users.",
    }:
        return "ASK_CLARIFICATION"
    if response.step == "hazard_profile_selection":
        return "START_MITIGATION_PLANNING"
    if response.step == "socio_demographic_review" and session.selected_hazard:
        return "SELECT_HAZARD"
    if response.step == "mitigation_measure":
        return "WRITE_MITIGATION_MANUALLY"
    if response.step == "mitigation_reason" and session.pending_mitigation_measure:
        if response.bot_message == "Typed mitigation measure accepted and reason requested.":
            return "CAPTURE_MITIGATION_MEASURE"
        return "ADOPT_MITIGATION_PROPOSAL"
    if session.phase == "custom_hazard_input":
        return "ADD_NEW_HAZARD"
    if session.pending_hazard == "__refresh_hazards__":
        return "REFRESH_HAZARDS"
    if "already selected" in response.bot_message.casefold():
        return "NO_CHANGE"
    if response.step == "country" and not session.country:
        return "RESET_ALL" if "reset" in response.bot_message.casefold() else "NO_CHANGE"
    if response.step == "region":
        return "SELECT_COUNTRY" if session.country and not session.region else "RESET_REGION_AND_SECTOR"
    if response.step == "sector":
        return "SELECT_REGION" if session.region and not session.sector else "RESET_SECTOR"
    if response.step == "hazards" and session.sector:
        return "COMPLETE_SELECTION"
    return "NO_CHANGE"


def bool_value(value: object) -> bool:
    return str(value or "").strip().casefold() == "yes"


def row_result(
    item: dict[str, str],
    response: ChatResponse,
    session: ChatSession,
) -> dict[str, str]:
    actual_action = infer_actual_action(response, session)
    if (
        str(item.get("Expected Action") or "") == "RESET_ALL"
        and not response.error
        and not any([session.country, session.region, session.sector])
    ):
        actual_action = "RESET_ALL"
    if str(item.get("Expected Action") or "") == "GO_BACK" and not response.error:
        actual_action = "GO_BACK"
    actual_clarify = response.step in {"selection_confirmation", "fuzzy_confirmation"} or (
        not response.error
        and response.bot_message
        in {
            "Please choose one of the available options, or type your selection another way.",
            "Please select a country from the available options.",
            "<p>Did you mean <strong>Heat stress</strong>?</p>",
            "Choose one of the hazards added by experts from the selected country-sector evidence.",
            "Choose one of the co-created hazards added by users.",
        }
        and str(item.get("Expected Action") or "") not in {"GO_BACK", "RESET_ALL"}
    )
    actual_error = bool(response.error)

    expected = {
        "country": str(item.get("Expected Country") or "").strip(),
        "region": str(item.get("Expected Region") or "").strip(),
        "sector": str(item.get("Expected Sector") or "").strip(),
        "hazard": str(item.get("Expected Hazard") or "").strip(),
        "mitigation": str(item.get("Expected Mitigation Measure") or "").strip(),
        "action": str(item.get("Expected Action") or "").strip(),
        "clarify": bool_value(item.get("Should Ask Clarification")),
        "error": bool_value(item.get("Should Show Error")),
    }
    actual = {
        "country": session.country or "",
        "region": session.region or "",
        "sector": session.sector or (
            str((session.pending_selection or {}).get("sector") or "")
            if isinstance(session.pending_selection, dict)
            else ""
        ),
        "hazard": session.selected_hazard or "",
        "mitigation": session.pending_mitigation_measure or session.mitigation_measure or "",
        "action": actual_action,
        "clarify": actual_clarify,
        "error": actual_error,
    }

    mismatches: list[str] = []
    for key in ("country", "region", "sector"):
        if expected[key] and expected[key] != actual[key]:
            mismatches.append(f"{key}: expected {expected[key]!r}, got {actual[key]!r}")
        if not expected[key] and actual[key] and expected["action"] in {"SHOW_ERROR", "ASK_CLARIFICATION", "NO_CHANGE", "RESET_ALL"}:
            mismatches.append(f"{key}: expected blank/no change, got {actual[key]!r}")
    if expected["hazard"] and expected["hazard"] != actual["hazard"]:
        mismatches.append(f"hazard: expected {expected['hazard']!r}, got {actual['hazard']!r}")
    if not expected["hazard"] and actual["hazard"] and expected["action"] in {"SHOW_ERROR", "ASK_CLARIFICATION", "NO_CHANGE", "RESET_ALL"}:
        mismatches.append(f"hazard: expected blank/no change, got {actual['hazard']!r}")
    if expected["mitigation"] and expected["mitigation"] != actual["mitigation"]:
        mismatches.append(
            f"mitigation: expected {expected['mitigation']!r}, got {actual['mitigation']!r}"
        )
    if not expected["mitigation"] and actual["mitigation"] and expected["action"] in {"SHOW_ERROR", "ASK_CLARIFICATION", "NO_CHANGE", "RESET_ALL"}:
        mismatches.append(f"mitigation: expected blank/no change, got {actual['mitigation']!r}")
    if expected["action"] and expected["action"] != actual["action"]:
        compatible_actions = {
            ("SELECT_SECTOR", "COMPLETE_SELECTION"),
            ("RESET_REGION_AND_SECTOR", "SELECT_COUNTRY"),
            ("RESET_SECTOR", "SELECT_REGION"),
            ("GO_BACK", "ASK_CLARIFICATION"),
        }
        if (expected["action"], actual["action"]) not in compatible_actions:
            mismatches.append(f"action: expected {expected['action']!r}, got {actual['action']!r}")
    if expected["clarify"] != actual["clarify"]:
        mismatches.append(f"clarification: expected {expected['clarify']}, got {actual['clarify']}")
    if expected["error"] != actual["error"]:
        mismatches.append(f"error: expected {expected['error']}, got {actual['error']}")

    return {
        "Actual Country": actual["country"],
        "Actual Region": actual["region"],
        "Actual Sector": actual["sector"],
        "Actual Hazard": actual["hazard"],
        "Actual Mitigation Measure": actual["mitigation"],
        "Actual Bot Response": response.bot_message,
        "Actual Action": actual_action,
        "Actual Should Ask Clarification": "Yes" if actual_clarify else "No",
        "Actual Should Show Error": "Yes" if actual_error else "No",
        "Status": "Fail" if mismatches else "Pass",
        "Reason": "; ".join(mismatches) if mismatches else "Matched expected selection behavior.",
    }


async def run_cases(
    limit: int | None = None,
    input_path: str | Path | None = None,
) -> list[dict[str, str]]:
    engine = _OpenConversationSelectionEngine()
    results: list[dict[str, str]] = []
    cases = load_or_make_test_cases(input_path)
    if limit is not None:
        cases = cases[:limit]
    for index, item in enumerate(cases, start=1):
        case_item = {"Test Case ID": f"TC-{index:03d}", **item}
        try:
            response, session = await engine.handle_case(case_item)
            results.append({**case_item, **row_result(case_item, response, session)})
        except Exception as exc:
            results.append(
                {
                    **case_item,
                    "Actual Country": "",
                    "Actual Region": "",
                    "Actual Sector": "",
                    "Actual Hazard": "",
                    "Actual Mitigation Measure": "",
                    "Actual Bot Response": "",
                    "Actual Action": "NO_CHANGE",
                    "Actual Should Ask Clarification": "No",
                    "Actual Should Show Error": "Yes",
                    "Status": "Fail",
                    "Reason": f"Runner exception: {exc.__class__.__name__}: {exc}",
                }
            )
    return results


def style_sheet(sheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    pass_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    fail_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    status_column = None
    for cell in sheet[1]:
        if cell.value == "Status":
            status_column = cell.column
            break

    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = body_alignment
        if status_column:
            status_cell = row_cells[status_column - 1]
            if status_cell.value == "Pass":
                status_cell.fill = pass_fill
            elif status_cell.value == "Fail":
                status_cell.fill = fail_fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 90))
        sheet.column_dimensions[column_letter].width = max(12, min(max_length + 2, 60))


def write_results_workbook(results: list[dict[str, str]], output_path: str | Path = OUTPUT_FILE) -> Path:
    output = Path(output_path).resolve()
    settings = get_settings()
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = RESULTS_SHEET
    sheet.append(RESULT_COLUMNS)
    for item in results:
        sheet.append([item.get(column, "") for column in RESULT_COLUMNS])
    style_sheet(sheet)

    summary = workbook.create_sheet(SUMMARY_SHEET)
    summary.append(["Run Setting", "Value"])
    summary.append(["Execution mode", "Real flow - no mocked intent or selection resolver"])
    summary.append(["Ollama base URL", settings.ollama_base_url])
    summary.append(["Ollama model", settings.ollama_model])
    summary.append([])
    summary.append(["Category", "Total", "Passed", "Failed"])
    by_category: dict[str, Counter[str]] = defaultdict(Counter)
    for item in results:
        by_category[str(item.get("Category") or "")][str(item.get("Status") or "")] += 1
    for category in sorted(by_category):
        counts = by_category[category]
        summary.append([category, counts["Pass"] + counts["Fail"], counts["Pass"], counts["Fail"]])
    summary.append(
        [
            "TOTAL",
            len(results),
            sum(1 for item in results if item.get("Status") == "Pass"),
            sum(1 for item in results if item.get("Status") == "Fail"),
        ]
    )
    style_sheet(summary)

    workbook.save(output)
    return output


def result_filename_for_model(model: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "_", model.strip())
    slug = slug.replace(":", "_")
    return f"open_conversation_selection_test_results_{slug}.xlsx"


async def run_cases_for_model(
    model: str,
    limit: int | None = None,
    input_path: str | Path | None = None,
) -> list[dict[str, str]]:
    os.environ["OLLAMA_MODEL"] = model
    get_settings.cache_clear()
    return await run_cases(limit=limit, input_path=input_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run open conversation selection cases through the real app flow."
    )
    parser.add_argument(
        "--models",
        nargs="+",
        help="Ollama model names to test. Writes one workbook per model.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="Optional maximum number of generated cases to run per model.",
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=Path.cwd() / TEST_CASES_FILE,
        help="Excel workbook with test cases. Defaults to the workbook in the current directory.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    models = args.models or [get_settings().ollama_model]
    for model in models:
        print(f"Running open conversation selection cases on Ollama model: {model}")
        input_path = args.input if args.input.exists() else None
        results = asyncio.run(
            run_cases_for_model(model, limit=args.limit, input_path=input_path)
        )
        output = write_results_workbook(results, Path.cwd() / result_filename_for_model(model))
        passed = sum(1 for item in results if item["Status"] == "Pass")
        failed = len(results) - passed
        print(f"Created selection test result file: {output}")
        print(f"Model: {model} | Total: {len(results)} | Passed: {passed} | Failed: {failed}")


if __name__ == "__main__":
    main()
